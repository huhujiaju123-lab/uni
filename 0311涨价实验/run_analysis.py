#!/usr/bin/env python3
"""
0311 价格实验分析脚本
分析期间: 2026-03-11 ~ 2026-03-16
"""

import json, time, os, subprocess, sys
from datetime import datetime

# ============ CONFIG ============
EXPERIMENT_START = '2026-03-11'
END_DATE = '2026-03-16'
GROUP_DT = '2026-03-15'
AUTH_FILE = os.path.expanduser('~/.claude/skills/cyberdata-query/auth.json')
OUTPUT_DIR = os.path.expanduser('~/Vibe coding/0311涨价实验')

# Experiment design
STRATEGIES = {
    '0-15天': {
        '涨价组1': '75折×2不限品',
        '涨价组2': '7折×2不限品',
        '对照组3': '6限品+7折不限品',
    },
    '16-30天': {
        '涨价组1': '5折+65折',
        '涨价组2': '5折+6折',
        '对照组3': '4折',
    },
    '31+天': {
        '涨价组1': '4折+$2.99+5折',
        '涨价组2': '4折+$2.99+5折',
        '对照组3': '3折+$2.99+5折',
    },
}

GROUP_ORDER = ['涨价组1', '涨价组2', '对照组3']
LIFECYCLE_ORDER = ['0-15天', '16-30天', '31+天']

# ============ API ============
def load_auth():
    with open(AUTH_FILE) as f:
        return json.load(f)

def run_sql(sql, auth, wait=6, max_poll=4):
    """Submit SQL query and poll for results"""
    cookies = auth['cookies']
    jwttoken = auth['jwttoken']
    ts = str(int(time.time() * 1000))

    submit_body = json.dumps({
        "_t": int(ts), "tenantId": "1001", "userId": "47",
        "projectId": "1906904360294313985", "resourceGroupId": 1,
        "taskId": "1990991087752757249", "variables": {},
        "sqlStatement": sql, "env": 5
    })

    r = subprocess.run(
        ['curl', '-s', 'https://idpcd.luckincoffee.us/api/dev/task/run',
         '-H', 'accept: application/json, text/plain, */*',
         '-H', 'content-type: application/json; charset=UTF-8',
         '-b', cookies, '-H', f'jwttoken: {jwttoken}',
         '-H', 'productkey: CyberData',
         '-H', 'origin: https://idpcd.luckincoffee.us',
         '--data-raw', submit_body],
        capture_output=True, text=True, timeout=30
    )

    resp = json.loads(r.stdout)
    if resp.get('code') != '200':
        raise Exception(f"Submit failed: {resp}")

    task_id = resp['data']

    for attempt in range(max_poll):
        time.sleep(wait)
        ts = str(int(time.time() * 1000))
        get_body = json.dumps({
            "_t": int(ts), "tenantId": "1001", "userId": "47",
            "projectId": "1906904360294313985", "env": 5,
            "taskInstanceId": task_id
        })

        r = subprocess.run(
            ['curl', '-s', 'https://idpcd.luckincoffee.us/api/logger/getQueryLog',
             '-H', 'accept: application/json, text/plain, */*',
             '-H', 'content-type: application/json; charset=UTF-8',
             '-b', cookies, '-H', f'jwttoken: {jwttoken}',
             '-H', 'productkey: CyberData',
             '-H', 'origin: https://idpcd.luckincoffee.us',
             '--data-raw', get_body],
            capture_output=True, text=True, timeout=30
        )

        resp = json.loads(r.stdout)
        if resp.get('code') == '200' and resp.get('data'):
            columns = resp['data'][0].get('columns', [])
            if columns and len(columns) >= 1:
                headers = columns[0]
                rows = columns[1:] if len(columns) > 1 else []
                return headers, rows

    raise Exception(f"Query timed out after {max_poll} polls (task: {task_id})")

def query(name, sql, auth, wait=6, max_poll=4):
    """Run a named query with logging"""
    print(f"\n{'='*50}")
    print(f"[{name}] Running...")
    start = time.time()
    headers, rows = run_sql(sql, auth, wait=wait, max_poll=max_poll)
    elapsed = time.time() - start
    print(f"[{name}] Done in {elapsed:.1f}s — {len(rows)} rows")
    if rows:
        print(f"  Headers: {headers}")
        for row in rows[:3]:
            print(f"  {row}")
        if len(rows) > 3:
            print(f"  ... and {len(rows)-3} more rows")
    return headers, rows

# ============ SQL QUERIES ============

SQL_OVERALL = f"""
WITH ab_users AS (
  SELECT DISTINCT user_no,
    CASE
      WHEN group_name = '0311涨价组1_7.6W' THEN '涨价组1'
      WHEN group_name = '0311涨价组2_4.6W' THEN '涨价组2'
      WHEN group_name = '0311对照组3_2.5W' THEN '对照组3'
    END AS grp
  FROM dw_ads.ads_marketing_t_user_group_d_his
  WHERE tenant = 'LKUS' AND dt = '{GROUP_DT}'
    AND group_name IN ('0311涨价组1_7.6W', '0311涨价组2_4.6W', '0311对照组3_2.5W')
),
user_orders AS (
  SELECT user_no,
    COUNT(*) AS drink_cnt,
    SUM(pay_amount) AS drink_pay,
    SUM(origin_price) AS drink_origin,
    COUNT(DISTINCT order_id) AS order_cnt
  FROM dw_dwd.dwd_t_ord_order_item_d_inc
  WHERE tenant = 'LKUS' AND order_status = 90
    AND order_category = '门店订单' AND one_category_name = 'Drink'
    AND shop_name NOT IN ('NJ Test Kitchen', 'NJ Test Kitchen 2')
    AND dt BETWEEN '{EXPERIMENT_START}' AND '{END_DATE}'
  GROUP BY user_no
)
SELECT ab.grp,
  COUNT(DISTINCT ab.user_no) AS total_users,
  COUNT(DISTINCT uo.user_no) AS order_users,
  COALESCE(SUM(uo.drink_cnt), 0) AS drink_cnt,
  ROUND(COALESCE(SUM(uo.drink_pay), 0), 2) AS drink_pay,
  ROUND(COALESCE(SUM(uo.drink_origin), 0), 2) AS drink_origin,
  COALESCE(SUM(uo.order_cnt), 0) AS order_cnt,
  ROUND(COALESCE(SUM(uo.drink_pay), 0) / COUNT(DISTINCT ab.user_no), 4) AS itt_pay,
  ROUND(COALESCE(SUM(uo.drink_cnt), 0) * 1.0 / COUNT(DISTINCT ab.user_no), 4) AS itt_cups,
  ROUND(COALESCE(SUM(uo.drink_pay), 0) / NULLIF(COALESCE(SUM(uo.drink_cnt), 0), 0), 2) AS price_per_cup,
  ROUND(COALESCE(SUM(uo.drink_pay), 0) / NULLIF(COALESCE(SUM(uo.order_cnt), 0), 0), 2) AS aov,
  ROUND(COUNT(DISTINCT uo.user_no) * 100.0 / COUNT(DISTINCT ab.user_no), 2) AS conv_rate,
  ROUND(COALESCE(SUM(uo.drink_pay), 0) / NULLIF(COALESCE(SUM(uo.drink_origin), 0), 0) * 100, 2) AS discount_rate
FROM ab_users ab
LEFT JOIN user_orders uo ON ab.user_no = uo.user_no
GROUP BY ab.grp ORDER BY ab.grp
"""

# Build lifecycle group name lists
_lc_names = []
for lc in LIFECYCLE_ORDER:
    for grp in GROUP_ORDER:
        suffix = {'0-15天': '0_15', '16-30天': '16_30', '31+天': '31+'}[lc]
        prefix = {'涨价组1': '涨价组1', '涨价组2': '涨价组2', '对照组3': '对照组3'}[grp]
        _lc_names.append(f"'0311{prefix}_{suffix}交易用户'")
_lc_in_clause = ', '.join(_lc_names)

SQL_LIFECYCLE = f"""
WITH lifecycle_users AS (
  SELECT DISTINCT user_no, group_name,
    CASE
      WHEN group_name LIKE '%涨价组1%' THEN '涨价组1'
      WHEN group_name LIKE '%涨价组2%' THEN '涨价组2'
      WHEN group_name LIKE '%对照组3%' THEN '对照组3'
    END AS grp,
    CASE
      WHEN group_name LIKE '%0\\_15%' THEN '0-15天'
      WHEN group_name LIKE '%16\\_30%' THEN '16-30天'
      WHEN group_name LIKE '%31+%' THEN '31+天'
    END AS lifecycle
  FROM dw_ads.ads_marketing_t_user_group_d_his
  WHERE tenant = 'LKUS' AND dt = '{GROUP_DT}'
    AND group_name IN ({_lc_in_clause})
),
user_orders AS (
  SELECT user_no,
    COUNT(*) AS drink_cnt,
    SUM(pay_amount) AS drink_pay,
    SUM(origin_price) AS drink_origin,
    COUNT(DISTINCT order_id) AS order_cnt
  FROM dw_dwd.dwd_t_ord_order_item_d_inc
  WHERE tenant = 'LKUS' AND order_status = 90
    AND order_category = '门店订单' AND one_category_name = 'Drink'
    AND shop_name NOT IN ('NJ Test Kitchen', 'NJ Test Kitchen 2')
    AND dt BETWEEN '{EXPERIMENT_START}' AND '{END_DATE}'
  GROUP BY user_no
)
SELECT lu.lifecycle, lu.grp,
  COUNT(DISTINCT lu.user_no) AS total_users,
  COUNT(DISTINCT uo.user_no) AS order_users,
  COALESCE(SUM(uo.drink_cnt), 0) AS drink_cnt,
  ROUND(COALESCE(SUM(uo.drink_pay), 0), 2) AS drink_pay,
  ROUND(COALESCE(SUM(uo.drink_origin), 0), 2) AS drink_origin,
  COALESCE(SUM(uo.order_cnt), 0) AS order_cnt,
  ROUND(COALESCE(SUM(uo.drink_pay), 0) / COUNT(DISTINCT lu.user_no), 4) AS itt_pay,
  ROUND(COALESCE(SUM(uo.drink_cnt), 0) * 1.0 / COUNT(DISTINCT lu.user_no), 4) AS itt_cups,
  ROUND(COALESCE(SUM(uo.drink_pay), 0) / NULLIF(COALESCE(SUM(uo.drink_cnt), 0), 0), 2) AS price_per_cup,
  ROUND(COALESCE(SUM(uo.drink_pay), 0) / NULLIF(COALESCE(SUM(uo.order_cnt), 0), 0), 2) AS aov,
  ROUND(COUNT(DISTINCT uo.user_no) * 100.0 / COUNT(DISTINCT lu.user_no), 2) AS conv_rate,
  ROUND(COALESCE(SUM(uo.drink_pay), 0) / NULLIF(COALESCE(SUM(uo.drink_origin), 0), 0) * 100, 2) AS discount_rate
FROM lifecycle_users lu
LEFT JOIN user_orders uo ON lu.user_no = uo.user_no
GROUP BY lu.lifecycle, lu.grp
ORDER BY lu.lifecycle, lu.grp
"""

SQL_VISIT_OVERALL = f"""
WITH ab_users AS (
  SELECT DISTINCT user_no,
    CASE
      WHEN group_name = '0311涨价组1_7.6W' THEN '涨价组1'
      WHEN group_name = '0311涨价组2_4.6W' THEN '涨价组2'
      WHEN group_name = '0311对照组3_2.5W' THEN '对照组3'
    END AS grp
  FROM dw_ads.ads_marketing_t_user_group_d_his
  WHERE tenant = 'LKUS' AND dt = '{GROUP_DT}'
    AND group_name IN ('0311涨价组1_7.6W', '0311涨价组2_4.6W', '0311对照组3_2.5W')
),
user_visits AS (
  SELECT login_id AS user_no
  FROM dw_dwd.dwd_mg_log_detail_d_inc
  WHERE local_dt BETWEEN '{EXPERIMENT_START}' AND '{END_DATE}'
    AND ((platform IN (1, 2) AND event = '$AppStart') OR (platform = 3))
    AND login_id IS NOT NULL AND login_id != ''
  GROUP BY login_id
)
SELECT ab.grp,
  COUNT(DISTINCT ab.user_no) AS total_users,
  COUNT(DISTINCT uv.user_no) AS visit_users,
  ROUND(COUNT(DISTINCT uv.user_no) * 100.0 / COUNT(DISTINCT ab.user_no), 2) AS visit_rate
FROM ab_users ab
LEFT JOIN user_visits uv ON ab.user_no = uv.user_no
GROUP BY ab.grp ORDER BY ab.grp
"""

SQL_VISIT_LIFECYCLE = f"""
WITH lifecycle_users AS (
  SELECT DISTINCT user_no,
    CASE
      WHEN group_name LIKE '%涨价组1%' THEN '涨价组1'
      WHEN group_name LIKE '%涨价组2%' THEN '涨价组2'
      WHEN group_name LIKE '%对照组3%' THEN '对照组3'
    END AS grp,
    CASE
      WHEN group_name LIKE '%0\\_15%' THEN '0-15天'
      WHEN group_name LIKE '%16\\_30%' THEN '16-30天'
      WHEN group_name LIKE '%31+%' THEN '31+天'
    END AS lifecycle
  FROM dw_ads.ads_marketing_t_user_group_d_his
  WHERE tenant = 'LKUS' AND dt = '{GROUP_DT}'
    AND group_name IN ({_lc_in_clause})
),
user_visits AS (
  SELECT login_id AS user_no
  FROM dw_dwd.dwd_mg_log_detail_d_inc
  WHERE local_dt BETWEEN '{EXPERIMENT_START}' AND '{END_DATE}'
    AND ((platform IN (1, 2) AND event = '$AppStart') OR (platform = 3))
    AND login_id IS NOT NULL AND login_id != ''
  GROUP BY login_id
)
SELECT lu.lifecycle, lu.grp,
  COUNT(DISTINCT lu.user_no) AS total_users,
  COUNT(DISTINCT uv.user_no) AS visit_users,
  ROUND(COUNT(DISTINCT uv.user_no) * 100.0 / COUNT(DISTINCT lu.user_no), 2) AS visit_rate
FROM lifecycle_users lu
LEFT JOIN user_visits uv ON lu.user_no = uv.user_no
GROUP BY lu.lifecycle, lu.grp
ORDER BY lu.lifecycle, lu.grp
"""

SQL_DAILY = f"""
WITH ab_users AS (
  SELECT DISTINCT user_no,
    CASE
      WHEN group_name = '0311涨价组1_7.6W' THEN '涨价组1'
      WHEN group_name = '0311涨价组2_4.6W' THEN '涨价组2'
      WHEN group_name = '0311对照组3_2.5W' THEN '对照组3'
    END AS grp
  FROM dw_ads.ads_marketing_t_user_group_d_his
  WHERE tenant = 'LKUS' AND dt = '{GROUP_DT}'
    AND group_name IN ('0311涨价组1_7.6W', '0311涨价组2_4.6W', '0311对照组3_2.5W')
),
daily_orders AS (
  SELECT user_no, dt,
    COUNT(*) AS drink_cnt,
    SUM(pay_amount) AS drink_pay,
    COUNT(DISTINCT order_id) AS order_cnt
  FROM dw_dwd.dwd_t_ord_order_item_d_inc
  WHERE tenant = 'LKUS' AND order_status = 90
    AND order_category = '门店订单' AND one_category_name = 'Drink'
    AND shop_name NOT IN ('NJ Test Kitchen', 'NJ Test Kitchen 2')
    AND dt BETWEEN '{EXPERIMENT_START}' AND '{END_DATE}'
  GROUP BY user_no, dt
)
SELECT d.dt, ab.grp,
  COUNT(DISTINCT d.user_no) AS order_users,
  SUM(d.drink_cnt) AS drink_cnt,
  ROUND(SUM(d.drink_pay), 2) AS drink_pay,
  ROUND(SUM(d.drink_pay) / NULLIF(SUM(d.drink_cnt), 0), 2) AS price_per_cup,
  SUM(d.order_cnt) AS order_cnt
FROM ab_users ab
JOIN daily_orders d ON ab.user_no = d.user_no
GROUP BY d.dt, ab.grp
ORDER BY d.dt, ab.grp
"""

# ============ DATA PROCESSING ============

def rows_to_dicts(headers, rows):
    """Convert API result to list of dicts"""
    return [{h: r[i] for i, h in enumerate(headers)} for r in rows]

def safe_float(v, default=0.0):
    try:
        return float(v) if v and v != 'None' and v != 'NULL' else default
    except (ValueError, TypeError):
        return default

def safe_int(v, default=0):
    try:
        return int(float(v)) if v and v != 'None' and v != 'NULL' else default
    except (ValueError, TypeError):
        return default

def calc_diff_pct(test_val, ctrl_val):
    """Calculate percentage difference vs control"""
    if ctrl_val == 0:
        return None
    return (test_val - ctrl_val) / ctrl_val * 100

def calc_diff_pp(test_val, ctrl_val):
    """Calculate pp difference for rate metrics"""
    return test_val - ctrl_val

# ============ EXCEL GENERATION ============

def generate_excel(overall_data, lifecycle_data, visit_overall, visit_lifecycle, daily_data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    subheader_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    pos_font = Font(color='008000', bold=True)
    neg_font = Font(color='CC0000', bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row, max_col):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    def style_cell(ws, row, col, value=None, is_diff=False):
        cell = ws.cell(row=row, column=col)
        if value is not None:
            cell.value = value
        cell.alignment = center_align
        cell.border = thin_border
        if is_diff and isinstance(value, (int, float)):
            if value > 0:
                cell.font = pos_font
            elif value < 0:
                cell.font = neg_font

    # ====== Sheet 1: 实验设计 ======
    ws1 = wb.active
    ws1.title = '实验设计'

    ws1.merge_cells('A1:F1')
    ws1.cell(1, 1, '0311 价格实验设计').font = Font(bold=True, size=14)
    ws1.cell(2, 1, f'实验周期: {EXPERIMENT_START} ~ {END_DATE} (6天)')
    ws1.cell(3, 1, f'全量实验用户: 147,489 | 涨价组1: 76,157 | 涨价组2: 46,220 | 对照组3: 25,112')

    # Strategy table
    row = 5
    headers = ['生命周期', '涨价组1 策略', '涨价组2 策略', '对照组3 策略', '涨价组1 人数', '涨价组2 人数', '对照组3 人数']
    for i, h in enumerate(headers, 1):
        ws1.cell(row, i, h)
    style_header(ws1, row, len(headers))

    lc_users = {
        '0-15天': (7934, 11771, 2791),
        '16-30天': (3532, 8468, 1249),
        '31+天': (64419, 25876, 20963),
    }

    for lc in LIFECYCLE_ORDER:
        row += 1
        ws1.cell(row, 1, lc).border = thin_border
        for i, grp in enumerate(GROUP_ORDER, 2):
            style_cell(ws1, row, i, STRATEGIES[lc][grp])
        users = lc_users[lc]
        for i, u in enumerate(users, 5):
            style_cell(ws1, row, i, f'{u:,}')

    for col in range(1, 8):
        ws1.column_dimensions[chr(64+col)].width = 20

    # ====== Sheet 2: 整体分析 ======
    ws2 = wb.create_sheet('整体分析')

    # Merge visit data into overall
    visit_map = {}
    if visit_overall:
        for d in visit_overall:
            visit_map[d['grp']] = d

    # Get control data for diff calculation
    ctrl = None
    for d in overall_data:
        if d['grp'] == '对照组3':
            ctrl = d
            break

    row = 1
    ws2.merge_cells('A1:N1')
    ws2.cell(1, 1, '整体分析：涨价组 vs 对照组').font = Font(bold=True, size=14)
    ws2.cell(2, 1, f'统计周期: {EXPERIMENT_START} ~ {END_DATE} | 数据口径: 仅饮品实收, 门店订单, order_status=90')

    row = 4
    metrics_headers = [
        '实验组', '进组人数', '到访人数', '到访率%', '下单人数', '下单率%', '访购率%',
        '饮品杯量', '人均杯量(ITT)', '饮品实收$', '人均实收$(ITT)',
        '单杯实收$', 'AOV$', '折扣率%'
    ]
    for i, h in enumerate(metrics_headers, 1):
        ws2.cell(row, i, h)
    style_header(ws2, row, len(metrics_headers))

    for d in overall_data:
        row += 1
        grp = d['grp']
        total = safe_int(d['total_users'])
        order_users = safe_int(d['order_users'])
        visit_users = safe_int(visit_map.get(grp, {}).get('visit_users', 0))
        visit_rate = safe_float(visit_map.get(grp, {}).get('visit_rate', 0))
        conv_rate = safe_float(d['conv_rate'])
        visit_conv = round(order_users / visit_users * 100, 2) if visit_users > 0 else 0

        values = [
            grp, total, visit_users, visit_rate, order_users, conv_rate, visit_conv,
            safe_int(d['drink_cnt']), safe_float(d['itt_cups']),
            safe_float(d['drink_pay']), safe_float(d['itt_pay']),
            safe_float(d['price_per_cup']), safe_float(d['aov']),
            safe_float(d['discount_rate'])
        ]
        for i, v in enumerate(values, 1):
            style_cell(ws2, row, i, v)

    # Diff rows
    if ctrl:
        ctrl_total = safe_int(ctrl['total_users'])
        ctrl_visit = safe_int(visit_map.get('对照组3', {}).get('visit_users', 0))
        ctrl_visit_rate = safe_float(visit_map.get('对照组3', {}).get('visit_rate', 0))
        ctrl_conv = safe_float(ctrl['conv_rate'])
        ctrl_visit_conv = round(safe_int(ctrl['order_users']) / ctrl_visit * 100, 2) if ctrl_visit > 0 else 0

        for d in overall_data:
            if d['grp'] == '对照组3':
                continue
            row += 1
            grp = d['grp']
            visit_users = safe_int(visit_map.get(grp, {}).get('visit_users', 0))
            visit_rate = safe_float(visit_map.get(grp, {}).get('visit_rate', 0))
            conv = safe_float(d['conv_rate'])
            vc = round(safe_int(d['order_users']) / visit_users * 100, 2) if visit_users > 0 else 0

            diffs = [
                f'{grp} vs 对照', '', '',
                round(visit_rate - ctrl_visit_rate, 2),
                '', round(conv - ctrl_conv, 2),
                round(vc - ctrl_visit_conv, 2),
                '',
                round(calc_diff_pct(safe_float(d['itt_cups']), safe_float(ctrl['itt_cups'])), 2) if safe_float(ctrl['itt_cups']) else '',
                '',
                round(calc_diff_pct(safe_float(d['itt_pay']), safe_float(ctrl['itt_pay'])), 2) if safe_float(ctrl['itt_pay']) else '',
                round(calc_diff_pct(safe_float(d['price_per_cup']), safe_float(ctrl['price_per_cup'])), 2) if safe_float(ctrl['price_per_cup']) else '',
                round(calc_diff_pct(safe_float(d['aov']), safe_float(ctrl['aov'])), 2) if safe_float(ctrl['aov']) else '',
                round(safe_float(d['discount_rate']) - safe_float(ctrl['discount_rate']), 2),
            ]
            for i, v in enumerate(diffs, 1):
                style_cell(ws2, row, i, v, is_diff=(i >= 3 and isinstance(v, (int, float))))

    for col_letter in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']:
        ws2.column_dimensions[col_letter].width = 14
    ws2.column_dimensions['A'].width = 18

    # ====== Sheet 3: 分层分析 ======
    ws3 = wb.create_sheet('分层分析')
    ws3.merge_cells('A1:N1')
    ws3.cell(1, 1, '分层分析：按生命周期').font = Font(bold=True, size=14)

    # Merge visit lifecycle data
    visit_lc_map = {}
    if visit_lifecycle:
        for d in visit_lifecycle:
            key = (d.get('lifecycle', ''), d.get('grp', ''))
            visit_lc_map[key] = d

    row = 3
    for lc in LIFECYCLE_ORDER:
        ws3.merge_cells(f'A{row}:N{row}')
        ws3.cell(row, 1, f'【{lc}】').font = Font(bold=True, size=12)
        ws3.cell(row, 1).fill = subheader_fill
        row += 1

        # Strategy info
        for grp in GROUP_ORDER:
            ws3.cell(row, 1, f'  {grp}: {STRATEGIES[lc][grp]}')
            row += 1

        # Headers
        for i, h in enumerate(metrics_headers, 1):
            ws3.cell(row, i, h)
        style_header(ws3, row, len(metrics_headers))

        # Find control for this lifecycle
        lc_ctrl = None
        lc_groups = []
        for d in lifecycle_data:
            if d.get('lifecycle') == lc:
                lc_groups.append(d)
                if d.get('grp') == '对照组3':
                    lc_ctrl = d

        for d in lc_groups:
            row += 1
            grp = d['grp']
            total = safe_int(d['total_users'])
            order_users = safe_int(d['order_users'])
            vk = (lc, grp)
            visit_users = safe_int(visit_lc_map.get(vk, {}).get('visit_users', 0))
            visit_rate = safe_float(visit_lc_map.get(vk, {}).get('visit_rate', 0))
            conv_rate = safe_float(d['conv_rate'])
            visit_conv = round(order_users / visit_users * 100, 2) if visit_users > 0 else 0

            values = [
                grp, total, visit_users, visit_rate, order_users, conv_rate, visit_conv,
                safe_int(d['drink_cnt']), safe_float(d['itt_cups']),
                safe_float(d['drink_pay']), safe_float(d['itt_pay']),
                safe_float(d['price_per_cup']), safe_float(d['aov']),
                safe_float(d['discount_rate'])
            ]
            for i, v in enumerate(values, 1):
                style_cell(ws3, row, i, v)

        # Diff rows
        if lc_ctrl:
            ctrl_visit_k = (lc, '对照组3')
            ctrl_visit_u = safe_int(visit_lc_map.get(ctrl_visit_k, {}).get('visit_users', 0))
            ctrl_visit_r = safe_float(visit_lc_map.get(ctrl_visit_k, {}).get('visit_rate', 0))
            ctrl_conv_r = safe_float(lc_ctrl['conv_rate'])
            ctrl_vc = round(safe_int(lc_ctrl['order_users']) / ctrl_visit_u * 100, 2) if ctrl_visit_u > 0 else 0

            for d in lc_groups:
                if d['grp'] == '对照组3':
                    continue
                row += 1
                grp = d['grp']
                vk = (lc, grp)
                vu = safe_int(visit_lc_map.get(vk, {}).get('visit_users', 0))
                vr = safe_float(visit_lc_map.get(vk, {}).get('visit_rate', 0))
                cr = safe_float(d['conv_rate'])
                vc = round(safe_int(d['order_users']) / vu * 100, 2) if vu > 0 else 0

                diffs = [
                    f'{grp} vs 对照', '', '',
                    round(vr - ctrl_visit_r, 2),
                    '', round(cr - ctrl_conv_r, 2),
                    round(vc - ctrl_vc, 2),
                    '',
                    round(calc_diff_pct(safe_float(d['itt_cups']), safe_float(lc_ctrl['itt_cups'])), 2) if safe_float(lc_ctrl['itt_cups']) else '',
                    '',
                    round(calc_diff_pct(safe_float(d['itt_pay']), safe_float(lc_ctrl['itt_pay'])), 2) if safe_float(lc_ctrl['itt_pay']) else '',
                    round(calc_diff_pct(safe_float(d['price_per_cup']), safe_float(lc_ctrl['price_per_cup'])), 2) if safe_float(lc_ctrl['price_per_cup']) else '',
                    round(calc_diff_pct(safe_float(d['aov']), safe_float(lc_ctrl['aov'])), 2) if safe_float(lc_ctrl['aov']) else '',
                    round(safe_float(d['discount_rate']) - safe_float(lc_ctrl['discount_rate']), 2),
                ]
                for i, v in enumerate(diffs, 1):
                    style_cell(ws3, row, i, v, is_diff=(i >= 3 and isinstance(v, (int, float))))

        row += 2  # Gap between lifecycle sections

    for col_letter in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']:
        ws3.column_dimensions[col_letter].width = 14
    ws3.column_dimensions['A'].width = 18

    # ====== Sheet 4: 日度趋势 ======
    ws4 = wb.create_sheet('日度趋势')
    ws4.merge_cells('A1:H1')
    ws4.cell(1, 1, '日度趋势').font = Font(bold=True, size=14)

    # Get total users per group for rate calculation
    total_map = {}
    for d in overall_data:
        total_map[d['grp']] = safe_int(d['total_users'])

    row = 3
    daily_headers = ['日期', '实验组', '下单人数', '下单率%', '饮品杯量', '人均杯量(ITT)', '饮品实收$', '单杯实收$']
    for i, h in enumerate(daily_headers, 1):
        ws4.cell(row, i, h)
    style_header(ws4, row, len(daily_headers))

    for d in daily_data:
        row += 1
        grp = d['grp']
        total = total_map.get(grp, 1)
        order_users = safe_int(d['order_users'])
        drink_cnt = safe_int(d['drink_cnt'])

        values = [
            d['dt'], grp, order_users,
            round(order_users / total * 100, 2),
            drink_cnt, round(drink_cnt / total, 4),
            safe_float(d['drink_pay']),
            safe_float(d['price_per_cup']),
        ]
        for i, v in enumerate(values, 1):
            style_cell(ws4, row, i, v)

    for col_letter in ['A','B','C','D','E','F','G','H']:
        ws4.column_dimensions[col_letter].width = 14

    # Save
    filepath = os.path.join(OUTPUT_DIR, '0311涨价实验分析.xlsx')
    wb.save(filepath)
    print(f"\nExcel saved: {filepath}")
    return filepath

# ============ HTML GENERATION ============

def generate_html(overall_data, lifecycle_data, visit_overall, visit_lifecycle, daily_data):
    """Generate interactive HTML report"""

    # Merge visit data
    visit_map = {}
    if visit_overall:
        for d in visit_overall:
            visit_map[d['grp']] = d

    visit_lc_map = {}
    if visit_lifecycle:
        for d in visit_lifecycle:
            key = (d.get('lifecycle', ''), d.get('grp', ''))
            visit_lc_map[key] = d

    # Get control
    ctrl = None
    for d in overall_data:
        if d['grp'] == '对照组3':
            ctrl = d
            break

    def diff_badge(val, fmt='pct', reverse=False):
        """Generate colored diff badge"""
        if val is None or val == '':
            return ''
        sign = '+' if val > 0 else ''
        color = '#e74c3c' if (val < 0 and not reverse) or (val > 0 and reverse) else '#27ae60'
        if abs(val) < 0.5:
            color = '#95a5a6'
        suffix = '%' if fmt == 'pct' else 'pp'
        return f'<span style="color:{color};font-weight:bold">{sign}{val:.2f}{suffix}</span>'

    def build_overall_table():
        rows_html = ''
        for d in overall_data:
            grp = d['grp']
            total = safe_int(d['total_users'])
            order_u = safe_int(d['order_users'])
            visit_u = safe_int(visit_map.get(grp, {}).get('visit_users', 0))
            visit_r = safe_float(visit_map.get(grp, {}).get('visit_rate', 0))
            conv = safe_float(d['conv_rate'])
            vc = round(order_u / visit_u * 100, 2) if visit_u > 0 else 0
            itt_pay = safe_float(d['itt_pay'])
            itt_cups = safe_float(d['itt_cups'])
            ppc = safe_float(d['price_per_cup'])
            aov = safe_float(d['aov'])
            disc = safe_float(d['discount_rate'])

            # Diff badges
            if ctrl and grp != '对照组3':
                c_visit_r = safe_float(visit_map.get('对照组3', {}).get('visit_rate', 0))
                c_conv = safe_float(ctrl['conv_rate'])
                c_vc_u = safe_int(visit_map.get('对照组3', {}).get('visit_users', 0))
                c_vc = round(safe_int(ctrl['order_users']) / c_vc_u * 100, 2) if c_vc_u > 0 else 0

                diff_itt = calc_diff_pct(itt_pay, safe_float(ctrl['itt_pay']))
                diff_cups = calc_diff_pct(itt_cups, safe_float(ctrl['itt_cups']))
                diff_ppc = calc_diff_pct(ppc, safe_float(ctrl['price_per_cup']))

                diff_html = f"""
                    <br>{diff_badge(visit_r - c_visit_r, 'pp')}
                    <br>{diff_badge(conv - c_conv, 'pp')}
                    <br>{diff_badge(vc - c_vc, 'pp')}
                    <br>{diff_badge(diff_cups)}
                    <br>{diff_badge(diff_itt)}
                    <br>{diff_badge(diff_ppc)}
                """
            else:
                diff_html = '<br>-<br>-<br>-<br>-<br>-<br>-'

            cls = 'ctrl' if grp == '对照组3' else ''
            rows_html += f"""
            <tr class="{cls}">
                <td><strong>{grp}</strong></td>
                <td>{total:,}</td>
                <td>{visit_u:,}<br><small>{visit_r:.2f}%</small></td>
                <td>{order_u:,}<br><small>{conv:.2f}%</small></td>
                <td>{vc:.2f}%</td>
                <td>{safe_int(d['drink_cnt']):,}<br><small>ITT: {itt_cups:.4f}</small></td>
                <td>${safe_float(d['drink_pay']):,.2f}<br><small>ITT: ${itt_pay:.4f}</small></td>
                <td>${ppc:.2f}</td>
                <td>${aov:.2f}</td>
                <td>{disc:.1f}%</td>
                <td>{diff_html}</td>
            </tr>"""
        return rows_html

    def build_lifecycle_section():
        html = ''
        for lc in LIFECYCLE_ORDER:
            lc_ctrl = None
            lc_groups = []
            for d in lifecycle_data:
                if d.get('lifecycle') == lc:
                    lc_groups.append(d)
                    if d.get('grp') == '对照组3':
                        lc_ctrl = d

            html += f'<h3>{lc}</h3>'
            html += '<div class="strategy-box">'
            for grp in GROUP_ORDER:
                html += f'<span class="strategy-tag">{grp}: {STRATEGIES[lc][grp]}</span>'
            html += '</div>'

            html += '''<table><thead><tr>
                <th>组</th><th>人数</th><th>到访(率)</th><th>下单(率)</th><th>访购率</th>
                <th>杯量(ITT)</th><th>实收(ITT)</th><th>单杯实收</th><th>折扣率</th><th>vs对照</th>
            </tr></thead><tbody>'''

            for d in lc_groups:
                grp = d['grp']
                total = safe_int(d['total_users'])
                order_u = safe_int(d['order_users'])
                vk = (lc, grp)
                visit_u = safe_int(visit_lc_map.get(vk, {}).get('visit_users', 0))
                visit_r = safe_float(visit_lc_map.get(vk, {}).get('visit_rate', 0))
                conv = safe_float(d['conv_rate'])
                vc = round(order_u / visit_u * 100, 2) if visit_u > 0 else 0
                itt_pay = safe_float(d['itt_pay'])
                itt_cups = safe_float(d['itt_cups'])
                ppc = safe_float(d['price_per_cup'])
                disc = safe_float(d['discount_rate'])

                if lc_ctrl and grp != '对照组3':
                    ck = (lc, '对照组3')
                    c_vr = safe_float(visit_lc_map.get(ck, {}).get('visit_rate', 0))
                    c_conv = safe_float(lc_ctrl['conv_rate'])
                    c_vu = safe_int(visit_lc_map.get(ck, {}).get('visit_users', 0))
                    c_vc = round(safe_int(lc_ctrl['order_users']) / c_vu * 100, 2) if c_vu > 0 else 0
                    d_itt = calc_diff_pct(itt_pay, safe_float(lc_ctrl['itt_pay']))
                    d_cups = calc_diff_pct(itt_cups, safe_float(lc_ctrl['itt_cups']))
                    d_ppc = calc_diff_pct(ppc, safe_float(lc_ctrl['price_per_cup']))
                    diff_html = f"ITT实收:{diff_badge(d_itt)}<br>杯量:{diff_badge(d_cups)}<br>单杯:{diff_badge(d_ppc)}<br>下单率:{diff_badge(conv-c_conv, 'pp')}"
                else:
                    diff_html = '基准'

                cls = 'ctrl' if grp == '对照组3' else ''
                html += f"""<tr class="{cls}">
                    <td><strong>{grp}</strong></td><td>{total:,}</td>
                    <td>{visit_u:,}<br><small>{visit_r:.1f}%</small></td>
                    <td>{order_u:,}<br><small>{conv:.2f}%</small></td>
                    <td>{vc:.2f}%</td>
                    <td>{safe_int(d['drink_cnt']):,}<br><small>{itt_cups:.4f}</small></td>
                    <td>${safe_float(d['drink_pay']):,.2f}<br><small>${itt_pay:.4f}</small></td>
                    <td>${ppc:.2f}</td><td>{disc:.1f}%</td>
                    <td>{diff_html}</td>
                </tr>"""

            html += '</tbody></table>'
        return html

    def build_daily_chart_data():
        """Build Chart.js data for daily trends"""
        dates = sorted(set(d['dt'] for d in daily_data))
        datasets = {}
        total_map = {}
        for d in overall_data:
            total_map[d['grp']] = safe_int(d['total_users'])

        colors = {'涨价组1': '#e74c3c', '涨价组2': '#f39c12', '对照组3': '#3498db'}

        for grp in GROUP_ORDER:
            grp_data = {d['dt']: d for d in daily_data if d['grp'] == grp}
            total = total_map.get(grp, 1)
            datasets[grp] = {
                'conv': [round(safe_int(grp_data.get(dt, {}).get('order_users', 0)) / total * 100, 2) for dt in dates],
                'ppc': [safe_float(grp_data.get(dt, {}).get('price_per_cup', 0)) for dt in dates],
                'cups_itt': [round(safe_int(grp_data.get(dt, {}).get('drink_cnt', 0)) / total, 4) for dt in dates],
                'color': colors.get(grp, '#999'),
            }

        return dates, datasets

    dates, chart_datasets = build_daily_chart_data()

    # Build chart JS
    chart_labels = json.dumps([d[-5:] for d in dates])  # MM-DD format

    def make_chart_datasets(metric):
        ds = []
        for grp in GROUP_ORDER:
            ds.append({
                'label': grp,
                'data': chart_datasets[grp][metric],
                'borderColor': chart_datasets[grp]['color'],
                'backgroundColor': chart_datasets[grp]['color'] + '20',
                'tension': 0.3,
                'pointRadius': 4,
            })
        return json.dumps(ds)

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>0311 价格实验分析报告</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f6fa; color: #2c3e50; padding: 20px; }}
.container {{ max-width: 1400px; margin: 0 auto; }}
h1 {{ font-size: 24px; margin-bottom: 8px; }}
h2 {{ font-size: 18px; margin: 24px 0 12px; padding-bottom: 8px; border-bottom: 2px solid #3498db; }}
h3 {{ font-size: 15px; margin: 16px 0 8px; color: #2F5496; }}
.meta {{ color: #666; font-size: 13px; margin-bottom: 20px; }}
.cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 12px; margin-bottom: 24px; }}
.card {{ background: white; border-radius: 8px; padding: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
.card .label {{ font-size: 12px; color: #888; }}
.card .value {{ font-size: 22px; font-weight: bold; margin: 4px 0; }}
.card .diff {{ font-size: 13px; }}
table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.1); margin-bottom: 16px; font-size: 13px; }}
th {{ background: #2F5496; color: white; padding: 10px 8px; text-align: center; font-weight: 600; }}
td {{ padding: 8px; text-align: center; border-bottom: 1px solid #eee; }}
tr:hover {{ background: #f8f9ff; }}
tr.ctrl {{ background: #f0f7ff; }}
tr.ctrl:hover {{ background: #e3efff; }}
.strategy-box {{ display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 12px; }}
.strategy-tag {{ background: #e8f0fe; padding: 4px 12px; border-radius: 4px; font-size: 12px; }}
.chart-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin: 16px 0; }}
.chart-box {{ background: white; border-radius: 8px; padding: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
.chart-box h4 {{ font-size: 14px; margin-bottom: 8px; color: #333; }}
.note {{ background: #fff3cd; border-left: 4px solid #ffc107; padding: 12px; margin: 16px 0; font-size: 13px; border-radius: 4px; }}
</style>
</head>
<body>
<div class="container">
<h1>0311 价格实验分析报告</h1>
<p class="meta">实验周期: {EXPERIMENT_START} ~ {END_DATE} (6天) | 数据口径: 仅饮品实收 + 门店订单 | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>

<div class="note">
<strong>实验设计：</strong>3组 × 3层生命周期。涨价组1 (76,157人, 42%) / 涨价组2 (46,220人, 26%) / 对照组3 (25,112人, 14%)。
各生命周期层发放不同折扣力度券，涨价组折扣力度更低（价格更高）。
</div>

<h2>整体对比</h2>
<table>
<thead><tr>
<th>组</th><th>人数</th><th>到访(率)</th><th>下单(率)</th><th>访购率</th>
<th>杯量(ITT)</th><th>实收(ITT)</th><th>单杯实收</th><th>AOV</th><th>折扣率</th><th>vs对照组</th>
</tr></thead>
<tbody>
{build_overall_table()}
</tbody>
</table>

<h2>分层分析（按生命周期）</h2>
{build_lifecycle_section()}

<h2>日度趋势</h2>
<div class="chart-grid">
<div class="chart-box"><h4>日下单率(%)</h4><canvas id="chartConv"></canvas></div>
<div class="chart-box"><h4>日单杯实收($)</h4><canvas id="chartPPC"></canvas></div>
<div class="chart-box"><h4>日人均杯量(ITT)</h4><canvas id="chartCups"></canvas></div>
</div>

<script>
const labels = {chart_labels};
const chartOpts = {{responsive:true, plugins:{{legend:{{position:'bottom'}}}}, scales:{{y:{{beginAtZero:false}}}}}};
new Chart(document.getElementById('chartConv'), {{type:'line', data:{{labels, datasets:{make_chart_datasets('conv')}}}, options:chartOpts}});
new Chart(document.getElementById('chartPPC'), {{type:'line', data:{{labels, datasets:{make_chart_datasets('ppc')}}}, options:chartOpts}});
new Chart(document.getElementById('chartCups'), {{type:'line', data:{{labels, datasets:{make_chart_datasets('cups_itt')}}}, options:chartOpts}});
</script>

<p class="meta" style="margin-top:24px;text-align:center;">Generated by Claude Code | Data source: CyberData DWD</p>
</div>
</body>
</html>"""

    filepath = os.path.join(OUTPUT_DIR, '0311涨价实验分析.html')
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"HTML saved: {filepath}")
    return filepath

# ============ MAIN ============

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    auth = load_auth()

    print("=" * 60)
    print("0311 价格实验分析")
    print(f"分析期间: {EXPERIMENT_START} ~ {END_DATE}")
    print(f"人群 dt: {GROUP_DT}")
    print("=" * 60)

    # Run all queries
    _, overall_rows = query("Q1: 整体指标", SQL_OVERALL, auth, wait=8, max_poll=5)
    _, lifecycle_rows = query("Q2: 分层指标", SQL_LIFECYCLE, auth, wait=8, max_poll=5)
    _, visit_overall_rows = query("Q3: 整体到访", SQL_VISIT_OVERALL, auth, wait=8, max_poll=5)
    _, visit_lc_rows = query("Q4: 分层到访", SQL_VISIT_LIFECYCLE, auth, wait=8, max_poll=5)
    _, daily_rows = query("Q5: 日度趋势", SQL_DAILY, auth, wait=8, max_poll=5)

    # Parse results
    overall_headers = ['grp', 'total_users', 'order_users', 'drink_cnt', 'drink_pay', 'drink_origin', 'order_cnt', 'itt_pay', 'itt_cups', 'price_per_cup', 'aov', 'conv_rate', 'discount_rate']
    lifecycle_headers = ['lifecycle', 'grp', 'total_users', 'order_users', 'drink_cnt', 'drink_pay', 'drink_origin', 'order_cnt', 'itt_pay', 'itt_cups', 'price_per_cup', 'aov', 'conv_rate', 'discount_rate']
    visit_headers = ['grp', 'total_users', 'visit_users', 'visit_rate']
    visit_lc_headers = ['lifecycle', 'grp', 'total_users', 'visit_users', 'visit_rate']
    daily_headers = ['dt', 'grp', 'order_users', 'drink_cnt', 'drink_pay', 'price_per_cup', 'order_cnt']

    overall_data = [{h: r[i] for i, h in enumerate(overall_headers)} for r in overall_rows] if overall_rows else []
    lifecycle_data = [{h: r[i] for i, h in enumerate(lifecycle_headers)} for r in lifecycle_rows] if lifecycle_rows else []
    visit_overall = [{h: r[i] for i, h in enumerate(visit_headers)} for r in visit_overall_rows] if visit_overall_rows else []
    visit_lifecycle = [{h: r[i] for i, h in enumerate(visit_lc_headers)} for r in visit_lc_rows] if visit_lc_rows else []
    daily_data = [{h: r[i] for i, h in enumerate(daily_headers)} for r in daily_rows] if daily_rows else []

    print(f"\n{'='*60}")
    print("数据收集完成，生成报告...")

    # Generate outputs
    excel_path = generate_excel(overall_data, lifecycle_data, visit_overall, visit_lifecycle, daily_data)
    html_path = generate_html(overall_data, lifecycle_data, visit_overall, visit_lifecycle, daily_data)

    # Save raw data as JSON for debugging
    raw_data = {
        'overall': overall_data,
        'lifecycle': lifecycle_data,
        'visit_overall': visit_overall,
        'visit_lifecycle': visit_lifecycle,
        'daily': daily_data,
        'config': {
            'experiment_start': EXPERIMENT_START,
            'end_date': END_DATE,
            'group_dt': GROUP_DT,
        }
    }
    json_path = os.path.join(OUTPUT_DIR, 'raw_data.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(raw_data, f, ensure_ascii=False, indent=2)
    print(f"Raw data: {json_path}")

    print(f"\n{'='*60}")
    print("分析完成！")
    print(f"  Excel: {excel_path}")
    print(f"  HTML:  {html_path}")
    print(f"  Data:  {json_path}")

if __name__ == '__main__':
    main()
