#!/usr/bin/env python3
"""0311 价格实验 — 5 子群分析脚本"""
import json, time, os, subprocess, sys
from datetime import datetime

EXPERIMENT_START = '2026-03-11'
END_DATE = '2026-03-16'
GROUP_DT = '2026-03-15'
AUTH_FILE = os.path.expanduser('~/.claude/skills/cyberdata-query/auth.json')
OUTPUT_DIR = os.path.expanduser('~/Vibe coding/0311涨价实验')

SUB_GROUPS = {
    'A': {'label': 'A:原涨价组1→涨价组1', 'users': 38095, 'from': '涨价组1(75折系)', 'to': '涨价组1', 'change': '基本不变'},
    'B': {'label': 'B:原涨价组2→涨价组1', 'users': 38062, 'from': '涨价组2(7折系)', 'to': '涨价组1', 'change': '上调(7折→75折系)'},
    'C': {'label': 'C:原对照组3→涨价组2', 'users': 25190, 'from': '对照组3(6折系)', 'to': '涨价组2', 'change': '上调(6折→7折系)'},
    'D': {'label': 'D:原对照组3→对照组3', 'users': 25112, 'from': '对照组3(6折系)', 'to': '对照组3', 'change': '无变化'},
    'E': {'label': 'E:原未覆盖→涨价组2', 'users': 21030, 'from': '未覆盖(6折系)', 'to': '涨价组2', 'change': '首次进入实验(6折→7折)'},
}
GROUP_ORDER = ['A', 'B', 'C', 'D', 'E']
COMPARISONS = [
    ('A', 'D', '长期涨价效果'),
    ('B', 'A', '组内提价边际冲击'),
    ('C', 'D', '⭐金标准: 同源50/50涨价冲击'),
    ('E', 'D', '新用户涨价效果'),
]

# ============ API ============
def load_auth():
    with open(AUTH_FILE) as f:
        return json.load(f)

def run_sql(sql, auth, wait=8, max_poll=5):
    cookies, jwttoken = auth['cookies'], auth['jwttoken']
    ts = str(int(time.time() * 1000))
    body = json.dumps({"_t": int(ts), "tenantId": "1001", "userId": "47",
        "projectId": "1906904360294313985", "resourceGroupId": 1,
        "taskId": "1990991087752757249", "variables": {}, "sqlStatement": sql, "env": 5})
    r = subprocess.run(['curl', '-s', 'https://idpcd.luckincoffee.us/api/dev/task/run',
        '-H', 'accept: application/json, text/plain, */*', '-H', 'content-type: application/json; charset=UTF-8',
        '-b', cookies, '-H', f'jwttoken: {jwttoken}', '-H', 'productkey: CyberData',
        '-H', 'origin: https://idpcd.luckincoffee.us', '--data-raw', body],
        capture_output=True, text=True, timeout=30)
    resp = json.loads(r.stdout)
    if resp.get('code') != '200': raise Exception(f"Submit failed: {resp}")
    task_id = resp['data']
    for _ in range(max_poll):
        time.sleep(wait)
        ts = str(int(time.time() * 1000))
        gb = json.dumps({"_t": int(ts), "tenantId": "1001", "userId": "47",
            "projectId": "1906904360294313985", "env": 5, "taskInstanceId": task_id})
        r = subprocess.run(['curl', '-s', 'https://idpcd.luckincoffee.us/api/logger/getQueryLog',
            '-H', 'accept: application/json, text/plain, */*', '-H', 'content-type: application/json; charset=UTF-8',
            '-b', cookies, '-H', f'jwttoken: {jwttoken}', '-H', 'productkey: CyberData',
            '-H', 'origin: https://idpcd.luckincoffee.us', '--data-raw', gb],
            capture_output=True, text=True, timeout=30)
        resp = json.loads(r.stdout)
        if resp.get('code') == '200' and resp.get('data'):
            cols = resp['data'][0].get('columns', [])
            if cols and len(cols) >= 1:
                return cols[0], cols[1:] if len(cols) > 1 else []
    raise Exception(f"Timeout (task: {task_id})")

def query(name, sql, auth, wait=8):
    print(f"\n{'='*50}\n[{name}] Running...")
    t0 = time.time()
    h, rows = run_sql(sql, auth, wait=wait)
    print(f"[{name}] {len(rows)} rows in {time.time()-t0:.1f}s")
    for r in rows[:3]: print(f"  {r}")
    if len(rows) > 3: print(f"  ... +{len(rows)-3} more")
    return h, rows

# ============ SUB-GROUP CTE (reused across queries) ============
SUB_GROUP_CTE = f"""
g0212 AS (
  SELECT DISTINCT user_no,
    CASE
      WHEN group_name = '0212价格实验30%分流涨价组1' THEN 'R1'
      WHEN group_name = '0212价格实验30%分流涨价组2' THEN 'R2'
      WHEN group_name = '0212价格实验40%分流对照组3' THEN 'C3'
      WHEN group_name = '0212价格实验未覆盖用户' THEN 'UC'
    END AS old_grp
  FROM dw_ads.ads_marketing_t_user_group_d_his
  WHERE tenant = 'LKUS' AND dt = '{GROUP_DT}'
    AND group_name IN ('0212价格实验30%分流涨价组1','0212价格实验30%分流涨价组2','0212价格实验40%分流对照组3','0212价格实验未覆盖用户')
),
g0311 AS (
  SELECT DISTINCT user_no,
    CASE
      WHEN group_name = '0311涨价组1_7.6W' THEN 'N1'
      WHEN group_name = '0311涨价组2_4.6W' THEN 'N2'
      WHEN group_name = '0311对照组3_2.5W' THEN 'N3'
    END AS new_grp
  FROM dw_ads.ads_marketing_t_user_group_d_his
  WHERE tenant = 'LKUS' AND dt = '{GROUP_DT}'
    AND group_name IN ('0311涨价组1_7.6W','0311涨价组2_4.6W','0311对照组3_2.5W')
),
sub_groups AS (
  SELECT a.user_no,
    CASE
      WHEN a.old_grp = 'R1' AND b.new_grp = 'N1' THEN 'A'
      WHEN a.old_grp = 'R2' AND b.new_grp = 'N1' THEN 'B'
      WHEN a.old_grp = 'C3' AND b.new_grp = 'N2' THEN 'C'
      WHEN a.old_grp = 'C3' AND b.new_grp = 'N3' THEN 'D'
      WHEN a.old_grp = 'UC' AND b.new_grp = 'N2' THEN 'E'
    END AS sub_grp
  FROM g0212 a JOIN g0311 b ON a.user_no = b.user_no
)
"""

ORDER_CTE = f"""
user_orders AS (
  SELECT user_no, COUNT(*) AS drink_cnt, SUM(pay_amount) AS drink_pay,
    SUM(origin_price) AS drink_origin, COUNT(DISTINCT order_id) AS order_cnt
  FROM dw_dwd.dwd_t_ord_order_item_d_inc
  WHERE tenant = 'LKUS' AND order_status = 90
    AND order_category = '门店订单' AND one_category_name = 'Drink'
    AND shop_name NOT IN ('NJ Test Kitchen', 'NJ Test Kitchen 2')
    AND dt BETWEEN '{EXPERIMENT_START}' AND '{END_DATE}'
  GROUP BY user_no
)
"""

METRICS_SELECT = """
  COUNT(DISTINCT sg.user_no) AS total_users,
  COUNT(DISTINCT uo.user_no) AS order_users,
  COALESCE(SUM(uo.drink_cnt), 0) AS drink_cnt,
  ROUND(COALESCE(SUM(uo.drink_pay), 0), 2) AS drink_pay,
  ROUND(COALESCE(SUM(uo.drink_origin), 0), 2) AS drink_origin,
  COALESCE(SUM(uo.order_cnt), 0) AS order_cnt,
  ROUND(COALESCE(SUM(uo.drink_pay), 0) / COUNT(DISTINCT sg.user_no), 4) AS itt_pay,
  ROUND(COALESCE(SUM(uo.drink_cnt), 0) * 1.0 / COUNT(DISTINCT sg.user_no), 4) AS itt_cups,
  ROUND(COALESCE(SUM(uo.drink_pay), 0) / NULLIF(COALESCE(SUM(uo.drink_cnt), 0), 0), 2) AS price_per_cup,
  ROUND(COALESCE(SUM(uo.drink_pay), 0) / NULLIF(COALESCE(SUM(uo.order_cnt), 0), 0), 2) AS aov,
  ROUND(COUNT(DISTINCT uo.user_no) * 100.0 / COUNT(DISTINCT sg.user_no), 2) AS conv_rate,
  ROUND(COALESCE(SUM(uo.drink_pay), 0) / NULLIF(COALESCE(SUM(uo.drink_origin), 0), 0) * 100, 2) AS discount_rate
"""

SQL_OVERALL = f"""WITH {SUB_GROUP_CTE}, {ORDER_CTE}
SELECT sg.sub_grp, {METRICS_SELECT}
FROM sub_groups sg LEFT JOIN user_orders uo ON sg.user_no = uo.user_no
WHERE sg.sub_grp IS NOT NULL
GROUP BY sg.sub_grp ORDER BY sg.sub_grp"""

LIFECYCLE_CTE = f"""
lifecycle AS (
  SELECT DISTINCT user_no,
    CASE
      WHEN group_name LIKE '%0\\_15%' THEN '0-15天'
      WHEN group_name LIKE '%16\\_30%' THEN '16-30天'
      WHEN group_name LIKE '%31+%' THEN '31+天'
    END AS lifecycle
  FROM dw_ads.ads_marketing_t_user_group_d_his
  WHERE tenant = 'LKUS' AND dt = '{GROUP_DT}'
    AND group_name LIKE '0311%交易用户'
)
"""

SQL_LIFECYCLE = f"""WITH {SUB_GROUP_CTE}, {LIFECYCLE_CTE}, {ORDER_CTE}
SELECT sg.sub_grp, COALESCE(lc.lifecycle, '无分层') AS lifecycle, {METRICS_SELECT}
FROM sub_groups sg
LEFT JOIN lifecycle lc ON sg.user_no = lc.user_no
LEFT JOIN user_orders uo ON sg.user_no = uo.user_no
WHERE sg.sub_grp IS NOT NULL
GROUP BY sg.sub_grp, lifecycle ORDER BY sg.sub_grp, lifecycle"""

SQL_VISIT = f"""WITH {SUB_GROUP_CTE},
user_visits AS (
  SELECT login_id AS user_no FROM dw_dwd.dwd_mg_log_detail_d_inc
  WHERE local_dt BETWEEN '{EXPERIMENT_START}' AND '{END_DATE}'
    AND ((platform IN (1, 2) AND event = '$AppStart') OR (platform = 3))
    AND login_id IS NOT NULL AND login_id != ''
  GROUP BY login_id
)
SELECT sg.sub_grp,
  COUNT(DISTINCT sg.user_no) AS total_users,
  COUNT(DISTINCT uv.user_no) AS visit_users,
  ROUND(COUNT(DISTINCT uv.user_no) * 100.0 / COUNT(DISTINCT sg.user_no), 2) AS visit_rate
FROM sub_groups sg LEFT JOIN user_visits uv ON sg.user_no = uv.user_no
WHERE sg.sub_grp IS NOT NULL
GROUP BY sg.sub_grp ORDER BY sg.sub_grp"""

SQL_DAILY = f"""WITH {SUB_GROUP_CTE},
daily_orders AS (
  SELECT user_no, dt, COUNT(*) AS drink_cnt, SUM(pay_amount) AS drink_pay, COUNT(DISTINCT order_id) AS order_cnt
  FROM dw_dwd.dwd_t_ord_order_item_d_inc
  WHERE tenant = 'LKUS' AND order_status = 90 AND order_category = '门店订单' AND one_category_name = 'Drink'
    AND shop_name NOT IN ('NJ Test Kitchen', 'NJ Test Kitchen 2')
    AND dt BETWEEN '{EXPERIMENT_START}' AND '{END_DATE}'
  GROUP BY user_no, dt
)
SELECT d.dt, sg.sub_grp,
  COUNT(DISTINCT d.user_no) AS order_users, SUM(d.drink_cnt) AS drink_cnt,
  ROUND(SUM(d.drink_pay), 2) AS drink_pay,
  ROUND(SUM(d.drink_pay) / NULLIF(SUM(d.drink_cnt), 0), 2) AS price_per_cup
FROM sub_groups sg JOIN daily_orders d ON sg.user_no = d.user_no
WHERE sg.sub_grp IS NOT NULL
GROUP BY d.dt, sg.sub_grp ORDER BY d.dt, sg.sub_grp"""

# ============ HELPERS ============
def to_dicts(headers, rows):
    return [{h: r[i] for i, h in enumerate(headers)} for r in rows]

def sf(v, d=0.0):
    try: return float(v) if v and v not in ('None','NULL','') else d
    except: return d

def si(v, d=0):
    try: return int(float(v)) if v and v not in ('None','NULL','') else d
    except: return d

def diff_pct(a, b):
    return round((a - b) / b * 100, 2) if b else None

# ============ EXCEL ============
def generate_excel(overall, lifecycle, visits, daily):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    hfont = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    goldfill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    greenfill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    pos_font = Font(color='008000', bold=True)
    neg_font = Font(color='CC0000', bold=True)
    ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    def hdr(ws, r, cols):
        for i, c in enumerate(cols, 1):
            cell = ws.cell(r, i, c); cell.font = hfont; cell.fill = hfill; cell.alignment = ca; cell.border = tb

    def val(ws, r, c, v, diff=False):
        cell = ws.cell(r, c, v); cell.alignment = ca; cell.border = tb
        if diff and isinstance(v, (int, float)):
            cell.font = pos_font if v > 0 else neg_font if v < 0 else None

    # Visit map
    visit_map = {d['sub_grp']: d for d in visits} if visits else {}

    # ====== Sheet 1: 整体分析 ======
    ws = wb.active; ws.title = '5组整体分析'
    ws.merge_cells('A1:O1')
    ws.cell(1, 1, '0311 价格实验 — 5 子群整体分析').font = Font(bold=True, size=14)
    ws.cell(2, 1, f'统计周期: {EXPERIMENT_START} ~ {END_DATE} | 口径: 仅饮品实收, 门店订单')

    r = 4
    cols = ['子群', '人数', '0212→0311', '到访人数', '到访率%', '下单人数', '下单率%', '访购率%',
            '杯量', 'ITT杯量', '实收$', 'ITT实收$', '单杯实收$', 'AOV$', '折扣率%']
    hdr(ws, r, cols)

    ctrl = None
    for d in overall:
        if d['sub_grp'] == 'D': ctrl = d

    for d in overall:
        r += 1
        g = d['sub_grp']
        info = SUB_GROUPS.get(g, {})
        total = si(d['total_users']); ou = si(d['order_users'])
        vu = si(visit_map.get(g, {}).get('visit_users', 0))
        vr = sf(visit_map.get(g, {}).get('visit_rate', 0))
        cr = sf(d['conv_rate']); vc = round(ou/vu*100, 2) if vu else 0
        fill = goldfill if g == 'D' else greenfill if g == 'C' else None
        vals = [info.get('label',''), total, info.get('change',''),
                vu, vr, ou, cr, vc,
                si(d['drink_cnt']), sf(d['itt_cups']),
                sf(d['drink_pay']), sf(d['itt_pay']),
                sf(d['price_per_cup']), sf(d['aov']), sf(d['discount_rate'])]
        for i, v in enumerate(vals, 1):
            val(ws, r, i, v)
            if fill: ws.cell(r, i).fill = fill

    # Diff rows
    if ctrl:
        r += 1; ws.cell(r, 1, '').border = tb  # blank row
        for test_g, ctrl_g, desc in COMPARISONS:
            td = next((d for d in overall if d['sub_grp'] == test_g), None)
            cd = next((d for d in overall if d['sub_grp'] == ctrl_g), None)
            if not td or not cd: continue
            r += 1
            val(ws, r, 1, f'{test_g} vs {ctrl_g}: {desc}')
            for i in [2, 3, 4]: val(ws, r, i, '')
            tv = si(visit_map.get(test_g, {}).get('visit_users', 0))
            cv = si(visit_map.get(ctrl_g, {}).get('visit_users', 0))
            tvr = sf(visit_map.get(test_g, {}).get('visit_rate', 0))
            cvr = sf(visit_map.get(ctrl_g, {}).get('visit_rate', 0))
            tcr = sf(td['conv_rate']); ccr = sf(cd['conv_rate'])
            tvc = round(si(td['order_users'])/tv*100,2) if tv else 0
            cvc = round(si(cd['order_users'])/cv*100,2) if cv else 0
            diffs = ['', '', '', '', round(tvr-cvr, 2), '', round(tcr-ccr, 2), round(tvc-cvc, 2), '',
                     diff_pct(sf(td['itt_cups']), sf(cd['itt_cups'])), '',
                     diff_pct(sf(td['itt_pay']), sf(cd['itt_pay'])),
                     diff_pct(sf(td['price_per_cup']), sf(cd['price_per_cup'])),
                     diff_pct(sf(td['aov']), sf(cd['aov'])),
                     round(sf(td['discount_rate'])-sf(cd['discount_rate']), 2)]
            for i, v in enumerate(diffs, 1):
                val(ws, r, i, v, diff=(isinstance(v, (int, float))))

    for c in 'ABCDEFGHIJKLMNO':
        ws.column_dimensions[c].width = 14
    ws.column_dimensions['A'].width = 28; ws.column_dimensions['C'].width = 22

    # ====== Sheet 2: 分层分析 ======
    ws2 = wb.create_sheet('分层分析')
    ws2.merge_cells('A1:N1')
    ws2.cell(1, 1, '5 子群 × 生命周期分层').font = Font(bold=True, size=14)

    r = 3
    lc_cols = ['子群', '生命周期', '人数', '下单人数', '下单率%', '杯量', 'ITT杯量', '实收$', 'ITT实收$', '单杯实收$', 'AOV$', '折扣率%']
    hdr(ws2, r, lc_cols)

    for d in lifecycle:
        r += 1
        g = d.get('sub_grp', ''); lc = d.get('lifecycle', '')
        total = si(d['total_users']); ou = si(d['order_users'])
        cr = sf(d['conv_rate'])
        fill = goldfill if g == 'D' else greenfill if g == 'C' else None
        vals = [SUB_GROUPS.get(g, {}).get('label', g), lc, total, ou, cr,
                si(d['drink_cnt']), sf(d['itt_cups']),
                sf(d['drink_pay']), sf(d['itt_pay']),
                sf(d['price_per_cup']), sf(d['aov']), sf(d['discount_rate'])]
        for i, v in enumerate(vals, 1):
            val(ws2, r, i, v)
            if fill: ws2.cell(r, i).fill = fill

    for c in 'ABCDEFGHIJKLMN':
        ws2.column_dimensions[c].width = 14
    ws2.column_dimensions['A'].width = 28

    # ====== Sheet 3: 日度趋势 ======
    ws3 = wb.create_sheet('日度趋势')
    ws3.merge_cells('A1:G1')
    ws3.cell(1, 1, '日度趋势（5子群）').font = Font(bold=True, size=14)

    total_map = {d['sub_grp']: si(d['total_users']) for d in overall}
    r = 3
    d_cols = ['日期', '子群', '下单人数', '下单率%', '杯量', '实收$', '单杯实收$']
    hdr(ws3, r, d_cols)

    for d in daily:
        r += 1
        g = d['sub_grp']; t = total_map.get(g, 1); ou = si(d['order_users'])
        vals = [d['dt'], SUB_GROUPS.get(g, {}).get('label', g), ou, round(ou/t*100, 2),
                si(d['drink_cnt']), sf(d['drink_pay']), sf(d['price_per_cup'])]
        for i, v in enumerate(vals, 1):
            val(ws3, r, i, v)

    for c in 'ABCDEFG':
        ws3.column_dimensions[c].width = 16
    ws3.column_dimensions['B'].width = 28

    path = os.path.join(OUTPUT_DIR, '0311涨价实验_5子群分析.xlsx')
    wb.save(path)
    print(f"\nExcel: {path}")
    return path

# ============ MAIN ============
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    auth = load_auth()
    print(f"{'='*60}\n0311 价格实验 — 5 子群分析\n{EXPERIMENT_START} ~ {END_DATE}\n{'='*60}")

    h1, r1 = query("Q1: 5子群整体", SQL_OVERALL, auth, wait=10)
    h2, r2 = query("Q2: 5子群×生命周期", SQL_LIFECYCLE, auth, wait=10)
    h3, r3 = query("Q3: 5子群到访", SQL_VISIT, auth, wait=10)
    h4, r4 = query("Q4: 日度趋势", SQL_DAILY, auth, wait=10)

    oh = ['sub_grp','total_users','order_users','drink_cnt','drink_pay','drink_origin','order_cnt','itt_pay','itt_cups','price_per_cup','aov','conv_rate','discount_rate']
    lh = ['sub_grp','lifecycle'] + oh[1:]
    vh = ['sub_grp','total_users','visit_users','visit_rate']
    dh = ['dt','sub_grp','order_users','drink_cnt','drink_pay','price_per_cup']

    overall = to_dicts(oh, r1) if r1 else []
    lifecycle = to_dicts(lh, r2) if r2 else []
    visits = to_dicts(vh, r3) if r3 else []
    daily = to_dicts(dh, r4) if r4 else []

    # Save raw
    raw = {'overall': overall, 'lifecycle': lifecycle, 'visits': visits, 'daily': daily}
    jp = os.path.join(OUTPUT_DIR, 'raw_subgroup_data.json')
    with open(jp, 'w', encoding='utf-8') as f: json.dump(raw, f, ensure_ascii=False, indent=2)

    generate_excel(overall, lifecycle, visits, daily)
    print(f"\n{'='*60}\nDone! Raw data: {jp}")

if __name__ == '__main__':
    main()
