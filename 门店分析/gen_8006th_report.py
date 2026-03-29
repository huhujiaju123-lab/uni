from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "8006th 数据全景"

# 样式定义
title_font = Font(name="Arial", size=16, bold=True)
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
section_font = Font(name="Arial", size=13, bold=True)
normal_font = Font(name="Arial", size=11)
note_font = Font(name="Arial", size=11, italic=True, color="666666")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_row(ws, row, data, font=normal_font, fill=None, alignment=None):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font
        cell.border = thin_border
        if fill:
            cell.fill = fill
        cell.alignment = alignment or (center if col > 1 else left)

row = 1

# 标题
ws.merge_cells("A1:C1")
cell = ws.cell(row=1, column=1, value="8006th 公寓合作券 — 数据全景")
cell.font = title_font
cell.alignment = Alignment(horizontal="left", vertical="center")
row = 3

# === 一、总体规模 ===
ws.merge_cells(f"A{row}:C{row}")
ws.cell(row=row, column=1, value="一、总体规模（2025-09 至今）").font = section_font
row += 1

set_row(ws, row, ["指标", "数值"], font=header_font, fill=header_fill, alignment=center)
row += 1

summary_data = [
    ["累计发券", "174 张（87 个券包 × 2张/包）"],
    ["领券用户", "47 人"],
    ["已核销", "60 张（34.5%）"],
    ["已过期", "60 张"],
    ["券面额", "50% OFF（无固定金额），有效期 31 天"],
    ["月上限", "100 个券包 → 实际用量不到上限的 25%"],
]
for i, d in enumerate(summary_data):
    fill = light_fill if i % 2 == 0 else None
    set_row(ws, row, d, fill=fill)
    row += 1

row += 1

# === 二、月度趋势 ===
ws.merge_cells(f"A{row}:D{row}")
ws.cell(row=row, column=1, value="二、月度趋势（持续下降）").font = section_font
row += 1

set_row(ws, row, ["月份", "领券人数", "核销数", "核销率"], font=header_font, fill=header_fill, alignment=center)
row += 1

monthly_data = [
    ["2025-09（首月）", 25, 15, "30%"],
    ["2025-10", 17, 13, "38%"],
    ["2025-11", 18, 11, "31%"],
    ["2025-12", 9, 8, "44%"],
    ["2026-01", 8, 7, "44%"],
    ["2026-02", 6, 6, "50%"],
    ["2026-03（进行中）", 4, 0, "—"],
]
for i, d in enumerate(monthly_data):
    fill = light_fill if i % 2 == 0 else None
    set_row(ws, row, d, fill=fill)
    row += 1

# 趋势说明
row += 1
ws.merge_cells(f"A{row}:D{row}")
cell = ws.cell(row=row, column=1, value="趋势：领券人数从首月 25 人下降到现在 4 人/月，大幅萎缩。核销率反而上升，说明留下来的是真实使用者。")
cell.font = note_font
cell.alignment = wrap
row += 2

# === 三、用户画像 ===
ws.merge_cells(f"A{row}:C{row}")
ws.cell(row=row, column=1, value="三、用户画像").font = section_font
row += 1

set_row(ws, row, ["用户类型", "人数", "占比"], font=header_font, fill=header_fill, alignment=center)
row += 1

user_data = [
    ["老用户（领券前已下单）", 28, "60%"],
    ["新客 - 当日首单", 6, "13%"],
    ["新客 - 领券后首单", 6, "13%"],
    ["领券未下单", 7, "15%"],
]
for i, d in enumerate(user_data):
    fill = light_fill if i % 2 == 0 else None
    set_row(ws, row, d, fill=fill)
    row += 1

row += 1
notes = [
    "实际拉新 12 人（当日 + 领券后首单），占领券用户的 26%",
    "60% 是已有用户使用优惠",
    "下单用户人均 10.4 单、人均消费 $43.42（含领券前后所有订单）",
]
for note in notes:
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws.cell(row=row, column=1, value=f"• {note}")
    cell.font = note_font
    cell.alignment = wrap
    row += 1

row += 1

# === 四、结论与建议 ===
ws.merge_cells(f"A{row}:D{row}")
ws.cell(row=row, column=1, value="四、结论与建议").font = section_font
row += 1

ws.merge_cells(f"A{row}:D{row}")
ws.cell(row=row, column=1, value="现状判断：效果一般").font = Font(name="Arial", size=11, bold=True)
row += 1

conclusions = [
    "6个月只拉了 12 个新客，月均 2 人",
    "月度领券量远低于 100 包上限，公寓推广力度不足",
    "多数是老用户使用优惠券，拉新效率低",
]
for c in conclusions:
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value=f"• {c}").font = normal_font
    row += 1

row += 1
ws.merge_cells(f"A{row}:D{row}")
ws.cell(row=row, column=1, value="建议：").font = Font(name="Arial", size=11, bold=True)
row += 1

suggestions = [
    "确认合作状态 — 原定3个月（2025年9-11月），已超期运行6个月",
    "核算补贴成本 — 60 张 50% OFF 券的实际补贴金额",
    "如继续：要求公寓方提高推广频次/可见度（月领不到 10 人，住户知晓度低）",
    "考虑对老用户限制（如仅新注册用户可领），或评估 ROI 决定是否终止",
]
for s in suggestions:
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value=f"• {s}").font = normal_font
    row += 1

# 列宽
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 15

# 行高
for r in range(1, row + 1):
    ws.row_dimensions[r].height = 22

output = "/Users/xiaoxiao/Downloads/8006th_公寓合作券_数据全景.xlsx"
wb.save(output)
print(f"已保存: {output}")
