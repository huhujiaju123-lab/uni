from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "100maidenlane 数据全景"

# 样式
title_font = Font(name="Arial", size=16, bold=True)
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
section_font = Font(name="Arial", size=13, bold=True)
normal_font = Font(name="Arial", size=11)
note_font = Font(name="Arial", size=11, italic=True, color="666666")
bold_font = Font(name="Arial", size=11, bold=True)
red_font = Font(name="Arial", size=11, bold=True, color="CC0000")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_row(ws, row, data, font=normal_font, fill=None, alignment=None):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font
        cell.border = thin_border
        if fill:
            cell.fill = fill
        cell.alignment = alignment or (center if col > 1 else left)

row = 1

# 标题
ws.merge_cells("A1:D1")
cell = ws.cell(row=1, column=1, value="100maidenlane 公寓合作券 — 数据全景")
cell.font = title_font
cell.alignment = Alignment(horizontal="left", vertical="center")
row = 3

# === 一、总体规模 ===
ws.merge_cells(f"A{row}:D{row}")
ws.cell(row=row, column=1, value="一、总体规模（2025-09）").font = section_font
row += 1

set_row(ws, row, ["指标", "数值"], font=header_font, fill=header_fill, alignment=center)
row += 1

summary_data = [
    ["累计发券", "10 张（5 个券包 × 2张/包）"],
    ["领券用户", "5 人"],
    ["已核销", "0 张（0%）"],
    ["已过期", "10 张（全部过期）"],
    ["券面额", "50% OFF（无固定金额），有效期 31 天"],
    ["活跃时间", "仅 2025-09-09 ~ 09-24，之后再无人领取"],
]
for i, d in enumerate(summary_data):
    fill = light_fill if i % 2 == 0 else None
    set_row(ws, row, d, fill=fill)
    row += 1

row += 1

# === 二、领券用户明细 ===
ws.merge_cells(f"A{row}:D{row}")
ws.cell(row=row, column=1, value="二、领券用户明细（共 5 人）").font = section_font
row += 1

set_row(ws, row, ["用户", "领券日", "用户类型", "说明"], font=header_font, fill=header_fill, alignment=center)
row += 1

user_data = [
    ["用户1", "2025-09-09", "当日首单", "领券当天完成首单，券未核销"],
    ["用户2", "2025-09-09", "领券后首单", "次日首单，券未核销"],
    ["用户3", "2025-09-10", "未下单", "领了券没用，过期"],
    ["用户4", "2025-09-10", "老用户", "7月已下过单，券未核销"],
    ["用户5", "2025-09-24", "老用户", "7月已下过单，券未核销"],
]
for i, d in enumerate(user_data):
    fill = light_fill if i % 2 == 0 else None
    set_row(ws, row, d, fill=fill)
    row += 1

row += 1

# === 三、用户画像汇总 ===
ws.merge_cells(f"A{row}:D{row}")
ws.cell(row=row, column=1, value="三、用户画像汇总").font = section_font
row += 1

set_row(ws, row, ["用户类型", "人数", "占比"], font=header_font, fill=header_fill, alignment=center)
row += 1

profile_data = [
    ["新客（当日首单 + 领券后首单）", 2, "40%"],
    ["老用户（领券前已下单）", 2, "40%"],
    ["领券未下单", 1, "20%"],
]
for i, d in enumerate(profile_data):
    fill = light_fill if i % 2 == 0 else None
    set_row(ws, row, d, fill=fill)
    row += 1

row += 1

# === 四、结论 ===
ws.merge_cells(f"A{row}:D{row}")
ws.cell(row=row, column=1, value="四、结论").font = section_font
row += 1

ws.merge_cells(f"A{row}:D{row}")
cell = ws.cell(row=row, column=1, value="现状判断：合作未实际落地，码已失效")
cell.font = red_font
row += 1

conclusions = [
    "总计仅 5 人领券，10 张券全部过期未核销，核销率 0%",
    "活跃窗口仅 2025-09-09 ~ 09-24（16天），之后再无人领取",
    "拉新 2 人，但均未使用券消费",
    "80 Pine 当时尚未入住（2026年3月才 40+ 住户入住），推广无法触达",
    "100maidenlane 码可废弃，用新码 80pine 重新启动合作",
]
for c in conclusions:
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value=f"• {c}").font = normal_font
    ws.row_dimensions[row].height = 24
    row += 1

# 列宽
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 32

for r in range(1, row + 1):
    if ws.row_dimensions[r].height is None:
        ws.row_dimensions[r].height = 22

output = "/Users/xiaoxiao/Downloads/100maidenlane_公寓合作券_数据全景.xlsx"
wb.save(output)
print(f"已保存: {output}")
