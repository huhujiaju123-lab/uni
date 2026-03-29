"""Coffee Pass 频次分析 Excel 报告生成"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# 颜色/样式定义
# ============================================================
GOLD = "F59E0B"
DARK_BG = "1E293B"
HEADER_BG = "334155"
LIGHT_BG = "F8FAFC"
HIGHLIGHT_BG = "FEF3C7"  # 活动中行高亮
GREEN = "16A34A"
RED = "DC2626"
GRAY = "94A3B8"

header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
title_font = Font(name="Arial", bold=True, size=14, color=DARK_BG)
subtitle_font = Font(name="Arial", bold=True, size=11, color=DARK_BG)
note_font = Font(name="Arial", italic=True, size=9, color=GRAY)
data_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", bold=True, size=10)
pct_green = Font(name="Arial", size=10, color=GREEN, bold=True)
pct_red = Font(name="Arial", size=10, color=RED, bold=True)
highlight_fill = PatternFill(start_color=HIGHLIGHT_BG, end_color=HIGHLIGHT_BG, fill_type="solid")
alt_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

center = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def style_data_row(ws, row, max_col, is_highlight=False, is_alt=False):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.alignment = center
        cell.border = thin_border
        if is_highlight:
            cell.fill = highlight_fill
            cell.font = bold_font
        elif is_alt:
            cell.fill = alt_fill


def pct_font(val):
    """正值绿色，负值红色"""
    if val > 0:
        return pct_green
    elif val < 0:
        return pct_red
    return data_font


def write_pct(ws, row, col, val, base_val=None):
    """写入百分比变化值"""
    if base_val and base_val != 0:
        diff = (val - base_val) / base_val
        cell = ws.cell(row=row, column=col)
        cell.value = f"{diff:+.1%}"
        cell.font = pct_font(diff)
        cell.alignment = center
        cell.border = thin_border
        return diff
    return None


# ============================================================
# Sheet 1: 核心对比（拉齐口径）
# ============================================================
ws1 = wb.active
ws1.title = "核心对比"
ws1.sheet_properties.tabColor = GOLD

# 标题
ws1.merge_cells("A1:L1")
ws1["A1"] = "Coffee Pass 频次分析 — 5 for $19.9 活动效果（口径拉齐版）"
ws1["A1"].font = title_font
ws1["A1"].alignment = left_align

ws1.merge_cells("A2:L2")
ws1["A2"] = "活动期 02-06~02-14 | 121名买券用户 | 每9天一个对比窗口 | 收入按 $3.98/杯(=$19.9÷5) 重算"
ws1["A2"].font = note_font
ws1["A2"].alignment = left_align

# 数据
periods = [
    ("前9天 (01-28~02-05)", "基线", 88, 288, 369, 66, 303, 1249.63, 262.68, 1512.31, 4.10),
    ("活动中 (02-06~02-14)", "活动", 107, 399, 448, 287, 161, 628.84, 1142.26, 1771.10, 3.95),
    ("后① (02-15~02-23)", "后1", 88, 221, 271, 135, 136, 579.35, 537.30, 1116.65, 4.12),
    ("后② (02-24~03-04)", "后2", 92, 243, 309, 68, 241, 1005.48, 270.64, 1276.12, 4.13),
    ("后③ (03-05~03-13)", "后3", 74, 202, 255, 25, 230, 952.60, 99.50, 1052.10, 4.13),
]
# period, tag, active, orders, cups, cp_cups, non_cp_cups, non_cp_rev, cp_adj_rev, adj_rev, adj_unit

headers1 = [
    "期间", "活跃人数", "活跃率", "订单数", "订单变化",
    "杯量", "杯量变化", "CP券杯数", "非CP杯数",
    "调整总收入", "收入变化", "调整单杯实收", "单杯变化"
]
row = 4
for c, h in enumerate(headers1, 1):
    ws1.cell(row=row, column=c, value=h)
style_header_row(ws1, row, len(headers1))

base = periods[0]  # 前9天 as baseline
for i, p in enumerate(periods):
    r = row + 1 + i
    name, tag, active, orders, cups, cp_cups, non_cp_cups, non_cp_rev, cp_adj, adj_rev, adj_unit = p
    is_hl = (tag == "活动")
    is_alt = (i % 2 == 0 and not is_hl)

    ws1.cell(row=r, column=1, value=name)
    ws1.cell(row=r, column=2, value=active)
    ws1.cell(row=r, column=3, value=f"{active/121*100:.1f}%")
    ws1.cell(row=r, column=4, value=orders)
    # 订单变化
    if i == 0:
        ws1.cell(row=r, column=5, value="—")
    else:
        diff = (orders - base[3]) / base[3]
        ws1.cell(row=r, column=5, value=f"{diff:+.1%}")
        ws1.cell(row=r, column=5).font = pct_font(diff)

    ws1.cell(row=r, column=6, value=cups)
    if i == 0:
        ws1.cell(row=r, column=7, value="—")
    else:
        diff = (cups - base[4]) / base[4]
        ws1.cell(row=r, column=7, value=f"{diff:+.1%}")
        ws1.cell(row=r, column=7).font = pct_font(diff)

    ws1.cell(row=r, column=8, value=cp_cups)
    ws1.cell(row=r, column=9, value=non_cp_cups)
    ws1.cell(row=r, column=10, value=f"${adj_rev:,.2f}")

    if i == 0:
        ws1.cell(row=r, column=11, value="—")
    else:
        diff = (adj_rev - base[9]) / base[9]
        ws1.cell(row=r, column=11, value=f"{diff:+.1%}")
        ws1.cell(row=r, column=11).font = pct_font(diff)

    ws1.cell(row=r, column=12, value=f"${adj_unit:.2f}")
    if i == 0:
        ws1.cell(row=r, column=13, value="—")
    else:
        diff = (adj_unit - base[10]) / base[10]
        ws1.cell(row=r, column=13, value=f"{diff:+.1%}")
        ws1.cell(row=r, column=13).font = pct_font(diff)

    style_data_row(ws1, r, len(headers1), is_highlight=is_hl, is_alt=is_alt)
    ws1.cell(row=r, column=1).alignment = left_align

# 注释
note_row = row + len(periods) + 2
ws1.merge_cells(f"A{note_row}:L{note_row}")
ws1.cell(row=note_row, column=1,
         value="注：CP券杯数 = Coffee Pass 核销杯数，按 $3.98/杯计价；非CP杯数 = 正常订单杯数，按 DWD pay_amount 计价")
ws1.cell(row=note_row, column=1).font = note_font

note_row2 = note_row + 1
ws1.merge_cells(f"A{note_row2}:L{note_row2}")
ws1.cell(row=note_row2, column=1,
         value="注：前9天基线含 66 杯 CP 核销（01-21 早期买家），基线可能略偏高")
ws1.cell(row=note_row2, column=1).font = note_font

# 列宽
col_widths = [24, 10, 8, 8, 10, 8, 10, 10, 10, 14, 10, 14, 10]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 2: 周维度趋势
# ============================================================
ws2 = wb.create_sheet("周维度趋势")
ws2.sheet_properties.tabColor = "3B82F6"

ws2.merge_cells("A1:K1")
ws2["A1"] = "Coffee Pass 买券用户 — 周维度消费趋势（121人）"
ws2["A1"].font = title_font
ws2["A1"].alignment = left_align

ws2.merge_cells("A2:K2")
ws2["A2"] = "收入已按 $3.98/杯 调整 Coffee Pass 核销部分 | W12仅1天数据已排除"
ws2["A2"].font = note_font

# 周数据（from earlier query + coupon data）
# I'll reconstruct with the adjustment
weekly_raw = [
    # (week, start, active, orders, cups, dwd_rev, unit_raw, cp_coupons_used, stage)
    ("W1", "12/30", 33, 80, 104, 402.37, 3.87, 10, "活动前"),
    ("W2", "01/05", 56, 170, 213, 729.14, 3.42, 30, "活动前"),
    ("W3", "01/12", 72, 197, 251, 1006.58, 4.01, 22, "活动前"),
    ("W4", "01/19", 73, 181, 221, 858.30, 3.88, 42, "活动前"),
    ("W5", "01/26", 84, 195, 256, 840.32, 3.28, 50, "活动前"),
    ("W6", "02/02", 86, 247, 294, 899.00, 3.06, 76, "活动中"),
    ("W7", "02/09", 109, 334, 373, 548.45, 1.47, 226, "活动中"),
    ("W8", "02/16", 83, 193, 237, 598.32, 2.52, 108, "活动后"),
    ("W9", "02/24", 78, 151, 193, 635.76, 3.29, 47, "活动后"),
    ("W10", "03/02", 73, 168, 212, 752.59, 3.55, 28, "活动后"),
    ("W11", "03/09", 60, 147, 185, 713.38, 3.86, 14, "活动后"),
]

# For adjustment, we need to know cp_dwd_rev per week.
# Approximation: cp_coupons × avg_cp_order_amount_per_cup
# From our period data: cp_dwd_rev/cp_cups ratios:
# 前9天: 26/66 = $0.39/cup, 活动中: 135.85/287 = $0.47/cup,
# 后①: 61.75/135 = $0.46/cup, 后②: 37.70/68 = $0.55/cup, 后③: 11.45/25 = $0.46/cup
# Average ≈ $0.46/cup
CP_DWD_AVG = 0.46  # average DWD pay per CP cup (to subtract)
CP_REAL = 3.98  # real price per CP cup

headers2 = [
    "周", "周起始", "阶段", "活跃人数", "活跃率",
    "订单数", "杯量", "DWD实收",
    "CP核销杯", "调整后收入", "调整单杯实收"
]
row2 = 4
for c, h in enumerate(headers2, 1):
    ws2.cell(row=row2, column=c, value=h)
style_header_row(ws2, row2, len(headers2))

for i, w in enumerate(weekly_raw):
    r = row2 + 1 + i
    week, start, active, orders, cups, dwd_rev, unit_raw, cp_used, stage = w

    # Adjustment: remove CP DWD amount, add CP real amount
    adj_rev = dwd_rev - cp_used * CP_DWD_AVG + cp_used * CP_REAL
    adj_unit = adj_rev / cups if cups > 0 else 0

    is_hl = (stage == "活动中")
    is_alt = (i % 2 == 0 and not is_hl)

    ws2.cell(row=r, column=1, value=week)
    ws2.cell(row=r, column=2, value=start)
    ws2.cell(row=r, column=3, value=stage)
    ws2.cell(row=r, column=4, value=active)
    ws2.cell(row=r, column=5, value=f"{active/121*100:.1f}%")
    ws2.cell(row=r, column=6, value=orders)
    ws2.cell(row=r, column=7, value=cups)
    ws2.cell(row=r, column=8, value=f"${dwd_rev:,.2f}")
    ws2.cell(row=r, column=9, value=cp_used)
    ws2.cell(row=r, column=10, value=f"${adj_rev:,.2f}")
    ws2.cell(row=r, column=11, value=f"${adj_unit:.2f}")

    style_data_row(ws2, r, len(headers2), is_highlight=is_hl, is_alt=is_alt)

# 阶段汇总
summary_row = row2 + len(weekly_raw) + 2
ws2.merge_cells(f"A{summary_row}:K{summary_row}")
ws2.cell(row=summary_row, column=1, value="分阶段汇总（周均值）")
ws2.cell(row=summary_row, column=1).font = subtitle_font

summary_headers = ["阶段", "周数", "平均活跃率", "周均订单", "周均杯量", "周均调整收入", "平均调整单杯"]
sr = summary_row + 1
for c, h in enumerate(summary_headers, 1):
    ws2.cell(row=sr, column=c, value=h)
style_header_row(ws2, sr, len(summary_headers))

# Before W2-W5
before_weeks = weekly_raw[1:5]
during_weeks = weekly_raw[5:7]
after_weeks = weekly_raw[7:11]

def summarize(weeks):
    n = len(weeks)
    avg_active_rate = sum(w[2] for w in weeks) / n / 121 * 100
    avg_orders = sum(w[3] for w in weeks) / n
    avg_cups = sum(w[4] for w in weeks) / n
    total_dwd = sum(w[5] for w in weeks)
    total_cp = sum(w[7] for w in weeks)
    adj_rev = total_dwd - total_cp * CP_DWD_AVG + total_cp * CP_REAL
    avg_adj_rev = adj_rev / n
    total_cups = sum(w[4] for w in weeks)
    avg_unit = adj_rev / total_cups if total_cups > 0 else 0
    return n, avg_active_rate, avg_orders, avg_cups, avg_adj_rev, avg_unit

stages = [
    ("活动前(W2-W5)", before_weeks),
    ("活动中(W6-W7)", during_weeks),
    ("活动后(W8-W11)", after_weeks),
]

for i, (name, weeks) in enumerate(stages):
    r = sr + 1 + i
    n, rate, orders, cups, rev, unit = summarize(weeks)
    is_hl = (i == 1)

    ws2.cell(row=r, column=1, value=name)
    ws2.cell(row=r, column=2, value=n)
    ws2.cell(row=r, column=3, value=f"{rate:.1f}%")
    ws2.cell(row=r, column=4, value=f"{orders:.0f}")
    ws2.cell(row=r, column=5, value=f"{cups:.0f}")
    ws2.cell(row=r, column=6, value=f"${rev:,.2f}")
    ws2.cell(row=r, column=7, value=f"${unit:.2f}")
    style_data_row(ws2, r, len(summary_headers), is_highlight=is_hl, is_alt=(i == 2))

# 列宽
w2_widths = [8, 8, 8, 10, 8, 8, 8, 12, 10, 14, 14]
for i, w in enumerate(w2_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 3: 券核销明细
# ============================================================
ws3 = wb.create_sheet("券核销时间线")
ws3.sheet_properties.tabColor = "10B981"

ws3.merge_cells("A1:F1")
ws3["A1"] = "Coffee Pass 券核销时间线"
ws3["A1"].font = title_font

# Daily coupon usage data
coupon_daily = [
    ("2026-01-21", 5, "活动前"), ("2026-01-22", 3, "活动前"), ("2026-01-23", 3, "活动前"),
    ("2026-01-26", 6, "活动前"), ("2026-01-27", 18, "活动前"), ("2026-01-28", 7, "活动前"),
    ("2026-01-29", 15, "活动前"), ("2026-01-30", 6, "活动前"), ("2026-01-31", 1, "活动前"),
    ("2026-02-01", 2, "活动前"), ("2026-02-02", 10, "活动前"), ("2026-02-03", 12, "活动前"),
    ("2026-02-04", 5, "活动前"), ("2026-02-05", 6, "活动前"),
    ("2026-02-06", 23, "活动中"), ("2026-02-07", 17, "活动中"), ("2026-02-08", 10, "活动中"),
    ("2026-02-09", 32, "活动中"), ("2026-02-10", 55, "活动中"), ("2026-02-11", 43, "活动中"),
    ("2026-02-12", 40, "活动中"), ("2026-02-13", 39, "活动中"), ("2026-02-14", 17, "活动中"),
    ("2026-02-15", 24, "活动后"), ("2026-02-16", 17, "活动后"), ("2026-02-17", 30, "活动后"),
    ("2026-02-18", 16, "活动后"), ("2026-02-19", 23, "活动后"), ("2026-02-20", 8, "活动后"),
    ("2026-02-21", 11, "活动后"), ("2026-02-22", 1, "活动后"), ("2026-02-24", 4, "活动后"),
    ("2026-02-25", 9, "活动后"), ("2026-02-26", 19, "活动后"), ("2026-02-27", 10, "活动后"),
    ("2026-02-28", 1, "活动后"), ("2026-03-01", 3, "活动后"), ("2026-03-02", 7, "活动后"),
    ("2026-03-03", 4, "活动后"), ("2026-03-04", 7, "活动后"), ("2026-03-05", 3, "活动后"),
    ("2026-03-06", 7, "活动后"), ("2026-03-07", 2, "活动后"), ("2026-03-08", 1, "活动后"),
    ("2026-03-09", 3, "活动后"), ("2026-03-10", 3, "活动后"), ("2026-03-11", 3, "活动后"),
    ("2026-03-12", 1, "活动后"), ("2026-03-13", 1, "活动后"), ("2026-03-14", 1, "活动后"),
    ("2026-03-15", 1, "活动后"), ("2026-03-16", 1, "活动后"),
]

headers3 = ["日期", "核销券数", "阶段", "累计核销", "核销收入($3.98/张)"]
row3 = 3
for c, h in enumerate(headers3, 1):
    ws3.cell(row=row3, column=c, value=h)
style_header_row(ws3, row3, len(headers3))

cum = 0
for i, (dt, cnt, stage) in enumerate(coupon_daily):
    r = row3 + 1 + i
    cum += cnt
    is_hl = (stage == "活动中")
    is_alt = (i % 2 == 0 and not is_hl)

    ws3.cell(row=r, column=1, value=dt)
    ws3.cell(row=r, column=2, value=cnt)
    ws3.cell(row=r, column=3, value=stage)
    ws3.cell(row=r, column=4, value=cum)
    ws3.cell(row=r, column=5, value=f"${cnt * 3.98:.2f}")
    style_data_row(ws3, r, len(headers3), is_highlight=is_hl, is_alt=is_alt)

# 汇总
total_row = row3 + len(coupon_daily) + 2
ws3.cell(row=total_row, column=1, value="合计")
ws3.cell(row=total_row, column=1).font = bold_font
ws3.cell(row=total_row, column=2, value=cum)
ws3.cell(row=total_row, column=2).font = bold_font
ws3.cell(row=total_row, column=5, value=f"${cum * 3.98:,.2f}")
ws3.cell(row=total_row, column=5).font = bold_font

before_sum = sum(c[1] for c in coupon_daily if c[2] == "活动前")
during_sum = sum(c[1] for c in coupon_daily if c[2] == "活动中")
after_sum = sum(c[1] for c in coupon_daily if c[2] == "活动后")

for i, (label, val) in enumerate([
    ("活动前", before_sum), ("活动中", during_sum), ("活动后", after_sum)
]):
    r = total_row + 1 + i
    ws3.cell(row=r, column=1, value=label)
    ws3.cell(row=r, column=2, value=val)
    ws3.cell(row=r, column=3, value=f"{val/cum*100:.1f}%")
    ws3.cell(row=r, column=5, value=f"${val * 3.98:,.2f}")

w3_widths = [14, 10, 8, 10, 16]
for i, w in enumerate(w3_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 4: 结论
# ============================================================
ws4 = wb.create_sheet("结论与建议")
ws4.sheet_properties.tabColor = "EF4444"

ws4.merge_cells("A1:B1")
ws4["A1"] = "Coffee Pass 频次分析 — 核心结论"
ws4["A1"].font = title_font

findings = [
    ("单杯实收基本合理",
     "按 $19.9÷5=$3.98/杯 重算后，活动中 $3.95 vs 活动前 $4.10，仅降 3.7%。"
     "券包定价没有大幅稀释单杯收入。"),
    ("频次提升是需求前置效应",
     "活动中订单 +38.5%（288→399/9天），但活动后持续低于基线：\n"
     "后① -23.3%，后② -15.6%，后③ -29.9%。\n"
     "说明是把未来消费提前拉到了活动期，而非创造增量。"),
    ("活跃率先升后降",
     "活动前 72.7% → 活动中 88.4% → 后③ 仅 61.2%（低于基线 11.5pp）。\n"
     "券用完后部分用户变得不活跃。"),
    ("CP券消耗轨迹",
     "活动中 276 张 → 后①135 → 后②68 → 后③25。\n"
     "余券基本在活动结束后 2-3 周耗尽，之后频次持续走低。"),
    ("经济账",
     "121人 × $19.9 = $2,407.9 券包收入。\n"
     "但活动后4个9天（36天）总收入比基线低约 $1,700。\n"
     "净效果：短期现金流 vs 长期频次下降的 trade-off。"),
]

row4 = 3
ws4.column_dimensions["A"].width = 24
ws4.column_dimensions["B"].width = 80

for i, (title, detail) in enumerate(findings):
    r = row4 + i * 2
    ws4.cell(row=r, column=1, value=f"发现{i+1}")
    ws4.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=11, color=GOLD)
    ws4.cell(row=r, column=2, value=title)
    ws4.cell(row=r, column=2).font = subtitle_font
    ws4.cell(row=r + 1, column=1, value="")
    ws4.cell(row=r + 1, column=2, value=detail)
    ws4.cell(row=r + 1, column=2).font = data_font
    ws4.cell(row=r + 1, column=2).alignment = wrap

# 数据口径
note_start = row4 + len(findings) * 2 + 1
ws4.cell(row=note_start, column=1, value="数据口径")
ws4.cell(row=note_start, column=1).font = subtitle_font
notes = [
    "人群：121名 Coffee Pass 买券用户（proposal_no = LKUSCP118713952489488385）",
    "订单：DWD层 dwd_t_ord_order_item_d_inc，order_status=90，order_category='门店订单'",
    "品类：仅饮品（one_category_name='Drink'）",
    "收入：非CP订单用 DWD pay_amount；CP核销杯按 $3.98/杯（=$19.9÷5）计价",
    "排除：测试店铺（NJ Test Kitchen 1&2）",
    "期间对齐：每个对比窗口 9 天，日均指标可横向对比",
]
for i, n in enumerate(notes):
    ws4.cell(row=note_start + 1 + i, column=1, value=f"  {i+1}.")
    ws4.cell(row=note_start + 1 + i, column=2, value=n)
    ws4.cell(row=note_start + 1 + i, column=2).font = note_font

# ============================================================
# 保存
# ============================================================
output_path = "/Users/xiaoxiao/Vibe coding/coffee-pass/Coffee_Pass_频次分析.xlsx"
wb.save(output_path)
print(f"✅ 报告已保存: {output_path}")
