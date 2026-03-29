"""Coffee Pass 频次分析 Excel 报告 v2 — 含券消费 vs 自然消费拆分"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# 样式
# ============================================================
GOLD = "F59E0B"
HEADER_BG = "334155"
LIGHT_BG = "F8FAFC"
HIGHLIGHT_BG = "FEF3C7"
GREEN = "16A34A"
RED = "DC2626"
GRAY = "94A3B8"
CP_BG = "EDE9FE"     # 券消费浅紫
ORG_BG = "DCFCE7"    # 自然消费浅绿

header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
title_font = Font(name="Arial", bold=True, size=14, color="1E293B")
subtitle_font = Font(name="Arial", bold=True, size=11, color="1E293B")
note_font = Font(name="Arial", italic=True, size=9, color=GRAY)
data_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", bold=True, size=10)
pct_green = Font(name="Arial", size=10, color=GREEN, bold=True)
pct_red = Font(name="Arial", size=10, color=RED, bold=True)
highlight_fill = PatternFill(start_color=HIGHLIGHT_BG, end_color=HIGHLIGHT_BG, fill_type="solid")
alt_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
cp_fill = PatternFill(start_color=CP_BG, end_color=CP_BG, fill_type="solid")
org_fill = PatternFill(start_color=ORG_BG, end_color=ORG_BG, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
center = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_row(ws, row, max_col, fill=None, font=None):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = thin_border
        cell.alignment = center
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        elif not cell.font or cell.font == Font():
            cell.font = data_font


def pct_f(val):
    return pct_green if val > 0.001 else (pct_red if val < -0.001 else data_font)


def diff_str(val, base):
    if base == 0:
        return "—"
    d = (val - base) / base
    return f"{d:+.1%}"


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# 原始数据
# ============================================================
# (period_name, cp_users, cp_orders, cp_cups, cp_dwd_rev,
#  org_users, org_orders, org_cups, org_dwd_rev, org_unit)
RAW = [
    ("前9天 (01-28~02-05)", 23, 64, 66, 26.00, 75, 224, 303, 1249.63, 4.12),
    ("活动中 (02-06~02-14)", 95, 276, 287, 135.85, 56, 123, 161, 628.84, 3.91),
    ("后A (02-15~02-23)", 71, 130, 135, 61.75, 37, 91, 136, 579.35, 4.26),
    ("后B (02-24~03-04)", 41, 64, 68, 37.70, 59, 179, 241, 1005.48, 4.17),
    ("后C (03-05~03-13)", 18, 24, 25, 11.45, 65, 178, 230, 952.60, 4.14),
]
CP_UNIT = 3.98
TOTAL_BUYERS = 121
BASELINE = RAW[0]  # 前9天

# ============================================================
# Sheet 1: 券消费 vs 自然消费拆分（核心）
# ============================================================
ws1 = wb.active
ws1.title = "券vs自然消费"
ws1.sheet_properties.tabColor = GOLD

ws1.merge_cells("A1:N1")
ws1["A1"] = "Coffee Pass 频次分析 — 券消费 vs 自然消费拆分（每期9天，口径拉齐）"
ws1["A1"].font = title_font
ws1["A1"].alignment = left_align

ws1.merge_cells("A2:N2")
ws1["A2"] = "121名买券用户 | 券方案 LKUSCP118713952489488385 | 活动期 02-06~02-14 | 券消费按 $3.98/杯计 | 仅饮品"
ws1["A2"].font = note_font

# Headers - 2 rows merged
r = 4
headers_top = [
    ("期间", 1, 1), ("总计", 2, 4), ("券消费(CP)", 5, 8), ("自然消费(Organic)", 9, 13), ("调整总收入", 14, 14)
]
for label, c_start, c_end in headers_top:
    if c_start == c_end:
        ws1.cell(row=r, column=c_start, value=label)
    else:
        ws1.merge_cells(start_row=r, start_column=c_start, end_row=r, end_column=c_end)
        ws1.cell(row=r, column=c_start, value=label)
style_row(ws1, r, 14, fill=header_fill, font=header_font)

r2 = 5
sub_headers = [
    "期间",
    "活跃人数", "总订单", "总杯量",
    "CP用户", "CP订单", "CP杯量", "CP收入($3.98)",
    "自然用户", "自然订单", "自然杯量", "自然变化", "自然单杯",
    "调整总收入"
]
for c, h in enumerate(sub_headers, 1):
    ws1.cell(row=r2, column=c, value=h)
style_row(ws1, r2, len(sub_headers), fill=header_fill, font=header_font)

# Data rows
for i, p in enumerate(RAW):
    r = 6 + i
    name, cp_u, cp_o, cp_c, cp_rev, org_u, org_o, org_c, org_rev, org_unit = p

    total_active = max(cp_u, org_u)  # approximate (some users in both)
    total_orders = cp_o + org_o
    total_cups = cp_c + org_c
    cp_adj_rev = cp_c * CP_UNIT
    adj_total = org_rev + cp_adj_rev
    org_diff = diff_str(org_c, BASELINE[7])  # baseline org_cups = 303

    is_hl = (i == 1)  # 活动中
    fill = highlight_fill if is_hl else (alt_fill if i % 2 == 0 else None)

    vals = [
        name,
        total_active, total_orders, total_cups,
        cp_u, cp_o, cp_c, f"${cp_adj_rev:,.2f}",
        org_u, org_o, org_c, org_diff, f"${org_unit:.2f}",
        f"${adj_total:,.2f}"
    ]
    for c, v in enumerate(vals, 1):
        ws1.cell(row=r, column=c, value=v)

    style_row(ws1, r, len(sub_headers), fill=fill)
    ws1.cell(row=r, column=1).alignment = left_align

    # Color the diff column
    if i > 0:
        d = (org_c - BASELINE[7]) / BASELINE[7]
        ws1.cell(row=r, column=12).font = pct_f(d)

    # Color CP columns with light purple, Organic with light green
    for c in range(5, 9):
        ws1.cell(row=r, column=c).fill = cp_fill if not is_hl else highlight_fill
    for c in range(9, 14):
        ws1.cell(row=r, column=c).fill = org_fill if not is_hl else highlight_fill

set_col_widths(ws1, [24, 10, 8, 8, 8, 8, 8, 14, 8, 8, 8, 10, 10, 14])

# Key insight box
note_r = 6 + len(RAW) + 1
ws1.merge_cells(f"A{note_r}:N{note_r}")
ws1.cell(row=note_r, column=1,
         value="KEY INSIGHT: 活动中自然消费杯量从 303 暴跌至 161 (-47%), "
               "说明 Coffee Pass 券替代了原有正价消费, 而非创造增量需求")
ws1.cell(row=note_r, column=1).font = Font(name="Arial", bold=True, size=10, color=RED)

note_r2 = note_r + 1
ws1.merge_cells(f"A{note_r2}:N{note_r2}")
ws1.cell(row=note_r2, column=1,
         value="活动后: 券耗完后(后C)自然消费仍低于基线 24%, 存在消费透支效应")
ws1.cell(row=note_r2, column=1).font = Font(name="Arial", bold=True, size=10, color="B45309")


# ============================================================
# Sheet 2: 综合对比（调整后口径）
# ============================================================
ws2 = wb.create_sheet("综合对比")
ws2.sheet_properties.tabColor = "3B82F6"

ws2.merge_cells("A1:L1")
ws2["A1"] = "Coffee Pass 综合对比 — 调整后收入 & 频次变化"
ws2["A1"].font = title_font
ws2["A1"].alignment = left_align

ws2.merge_cells("A2:L2")
ws2["A2"] = "CP核销杯按 $3.98/杯计价 | 每期9天拉齐 | vs 前9天基线"
ws2["A2"].font = note_font

headers2 = [
    "期间", "活跃人数", "活跃率",
    "总订单", "订单变化", "总杯量", "杯量变化",
    "CP杯量", "CP占比",
    "调整总收入", "收入变化", "调整单杯"
]
r = 4
for c, h in enumerate(headers2, 1):
    ws2.cell(row=r, column=c, value=h)
style_row(ws2, r, len(headers2), fill=header_fill, font=header_font)

for i, p in enumerate(RAW):
    r = 5 + i
    name, cp_u, cp_o, cp_c, cp_rev, org_u, org_o, org_c, org_rev, org_unit = p

    total_orders = cp_o + org_o
    total_cups = cp_c + org_c
    cp_adj_rev = cp_c * CP_UNIT
    adj_total = org_rev + cp_adj_rev
    adj_unit = adj_total / total_cups if total_cups > 0 else 0
    active = max(cp_u, org_u)

    b_orders = BASELINE[2] + BASELINE[6]  # 64 + 224 = 288
    b_cups = BASELINE[3] + BASELINE[7]    # 66 + 303 = 369
    b_adj = BASELINE[8] + BASELINE[3] * CP_UNIT  # 1249.63 + 66*3.98 = 1512.31

    is_hl = (i == 1)
    fill = highlight_fill if is_hl else (alt_fill if i % 2 == 0 else None)

    ws2.cell(row=r, column=1, value=name)
    ws2.cell(row=r, column=2, value=active)
    ws2.cell(row=r, column=3, value=f"{active/121*100:.1f}%")
    ws2.cell(row=r, column=4, value=total_orders)
    ws2.cell(row=r, column=5, value="—" if i == 0 else diff_str(total_orders, b_orders))
    ws2.cell(row=r, column=6, value=total_cups)
    ws2.cell(row=r, column=7, value="—" if i == 0 else diff_str(total_cups, b_cups))
    ws2.cell(row=r, column=8, value=cp_c)
    ws2.cell(row=r, column=9, value=f"{cp_c/total_cups*100:.0f}%" if total_cups > 0 else "0%")
    ws2.cell(row=r, column=10, value=f"${adj_total:,.2f}")
    ws2.cell(row=r, column=11, value="—" if i == 0 else diff_str(adj_total, b_adj))
    ws2.cell(row=r, column=12, value=f"${adj_unit:.2f}")

    style_row(ws2, r, len(headers2), fill=fill)
    ws2.cell(row=r, column=1).alignment = left_align

    # Color diffs
    if i > 0:
        for col_idx, val, base_val in [
            (5, total_orders, b_orders),
            (7, total_cups, b_cups),
            (11, adj_total, b_adj),
        ]:
            d = (val - base_val) / base_val
            ws2.cell(row=r, column=col_idx).font = pct_f(d)

set_col_widths(ws2, [24, 10, 8, 8, 10, 8, 10, 8, 8, 14, 10, 10])


# ============================================================
# Sheet 3: 自然消费趋势（最关键的 sheet）
# ============================================================
ws3 = wb.create_sheet("自然消费趋势")
ws3.sheet_properties.tabColor = "10B981"

ws3.merge_cells("A1:I1")
ws3["A1"] = "自然消费(去除券消费后)趋势 — Coffee Pass 的真实影响"
ws3["A1"].font = title_font
ws3["A1"].alignment = left_align

ws3.merge_cells("A2:I2")
ws3["A2"] = "仅看非 Coffee Pass 订单 | 这是判断活动是否创造增量的核心视角"
ws3["A2"].font = note_font

headers3 = [
    "期间", "自然用户", "自然用户变化", "自然订单", "自然订单变化",
    "自然杯量", "自然杯量变化", "自然单杯实收", "自然收入"
]
r = 4
for c, h in enumerate(headers3, 1):
    ws3.cell(row=r, column=c, value=h)
style_row(ws3, r, len(headers3), fill=header_fill, font=header_font)

b_org_u = BASELINE[5]   # 75
b_org_o = BASELINE[6]   # 224
b_org_c = BASELINE[7]   # 303
b_org_rev = BASELINE[8] # 1249.63

for i, p in enumerate(RAW):
    r = 5 + i
    name = p[0]
    org_u, org_o, org_c, org_rev, org_unit = p[5], p[6], p[7], p[8], p[9]

    is_hl = (i == 1)
    fill = highlight_fill if is_hl else (alt_fill if i % 2 == 0 else None)

    ws3.cell(row=r, column=1, value=name)
    ws3.cell(row=r, column=2, value=org_u)
    ws3.cell(row=r, column=3, value="—" if i == 0 else diff_str(org_u, b_org_u))
    ws3.cell(row=r, column=4, value=org_o)
    ws3.cell(row=r, column=5, value="—" if i == 0 else diff_str(org_o, b_org_o))
    ws3.cell(row=r, column=6, value=org_c)
    ws3.cell(row=r, column=7, value="—" if i == 0 else diff_str(org_c, b_org_c))
    ws3.cell(row=r, column=8, value=f"${org_unit:.2f}")
    ws3.cell(row=r, column=9, value=f"${org_rev:,.2f}")

    style_row(ws3, r, len(headers3), fill=fill)
    ws3.cell(row=r, column=1).alignment = left_align

    if i > 0:
        for col_idx, val, base_val in [
            (3, org_u, b_org_u), (5, org_o, b_org_o), (7, org_c, b_org_c)
        ]:
            d = (val - base_val) / base_val
            ws3.cell(row=r, column=col_idx).font = pct_f(d)

# Insight
note_r = 5 + len(RAW) + 1
insights = [
    ("活动中", "自然杯量 303 -> 161 (-47%): 券消费替代了正价消费, 非增量",
     RED),
    ("后A(02-15~02-23)", "自然杯量仅 136 (-55%): 余券仍在消耗, 自然购买几乎腰斩",
     RED),
    ("后B(02-24~03-04)", "自然杯量恢复至 241 (-20%): 券基本耗完, 开始恢复",
     "B45309"),
    ("后C(03-05~03-13)", "自然杯量 230 (-24%): 仍未回到基线, 存在消费透支效应",
     "B45309"),
    ("自然单杯实收", "始终稳定在 $3.91~$4.26, 价格没被影响, 问题在频次",
     "6B7280"),
]
for j, (label, text, color) in enumerate(insights):
    r = note_r + j
    ws3.cell(row=r, column=1, value=label)
    ws3.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=10, color=color)
    ws3.merge_cells(f"B{r}:I{r}")
    ws3.cell(row=r, column=2, value=text)
    ws3.cell(row=r, column=2).font = Font(name="Arial", size=10, color=color)

set_col_widths(ws3, [24, 10, 12, 10, 12, 10, 12, 12, 14])


# ============================================================
# Sheet 4: 券核销时间线
# ============================================================
ws4 = wb.create_sheet("券核销时间线")
ws4.sheet_properties.tabColor = "8B5CF6"

ws4.merge_cells("A1:G1")
ws4["A1"] = "Coffee Pass 券核销时间线（逐日）"
ws4["A1"].font = title_font

ws4.merge_cells("A2:G2")
ws4["A2"] = "券方案: LKUSCP118713952489488385 | 每张券 = $3.98 (=$19.9/5)"
ws4["A2"].font = note_font

coupon_daily = [
    ("2026-01-21", 5, "活动前"), ("2026-01-22", 3, "活动前"), ("2026-01-23", 3, "活动前"),
    ("2026-01-26", 6, "活动前"), ("2026-01-27", 18, "活动前"), ("2026-01-28", 7, "活动前"),
    ("2026-01-29", 15, "活动前"), ("2026-01-30", 6, "活动前"), ("2026-01-31", 1, "活动前"),
    ("2026-02-01", 2, "活动前"), ("2026-02-02", 10, "活动前"), ("2026-02-03", 12, "活动前"),
    ("2026-02-04", 5, "活动前"), ("2026-02-05", 6, "活动前"),
    ("2026-02-06", 23, "活动中"), ("2026-02-07", 17, "活动中"), ("2026-02-08", 10, "活动中"),
    ("2026-02-09", 32, "活动中"), ("2026-02-10", 55, "活动中"), ("2026-02-11", 43, "活动中"),
    ("2026-02-12", 40, "活动中"), ("2026-02-13", 39, "活动中"), ("2026-02-14", 17, "活动中"),
    ("2026-02-15", 24, "活动后"), ("2026-02-16", 17, "活动后"), ("2026-02-17", 30, "活动后"),
    ("2026-02-18", 16, "活动后"), ("2026-02-19", 23, "活动后"), ("2026-02-20", 8, "活动后"),
    ("2026-02-21", 11, "活动后"), ("2026-02-22", 1, "活动后"), ("2026-02-24", 4, "活动后"),
    ("2026-02-25", 9, "活动后"), ("2026-02-26", 19, "活动后"), ("2026-02-27", 10, "活动后"),
    ("2026-02-28", 1, "活动后"), ("2026-03-01", 3, "活动后"), ("2026-03-02", 7, "活动后"),
    ("2026-03-03", 4, "活动后"), ("2026-03-04", 7, "活动后"), ("2026-03-05", 3, "活动后"),
    ("2026-03-06", 7, "活动后"), ("2026-03-07", 2, "活动后"), ("2026-03-08", 1, "活动后"),
    ("2026-03-09", 3, "活动后"), ("2026-03-10", 3, "活动后"), ("2026-03-11", 3, "活动后"),
    ("2026-03-12", 1, "活动后"), ("2026-03-13", 1, "活动后"), ("2026-03-14", 1, "活动后"),
    ("2026-03-15", 1, "活动后"), ("2026-03-16", 1, "活动后"),
]

headers4 = ["日期", "核销券数", "阶段", "累计核销", "累计占总发券%", "单日收入", "累计收入"]
r = 4
for c, h in enumerate(headers4, 1):
    ws4.cell(row=r, column=c, value=h)
style_row(ws4, r, len(headers4), fill=header_fill, font=header_font)

TOTAL_COUPONS = 695
cum = 0
for i, (dt, cnt, stage) in enumerate(coupon_daily):
    r = 5 + i
    cum += cnt
    is_hl = (stage == "活动中")
    fill = highlight_fill if is_hl else (alt_fill if i % 2 == 0 else None)

    ws4.cell(row=r, column=1, value=dt)
    ws4.cell(row=r, column=2, value=cnt)
    ws4.cell(row=r, column=3, value=stage)
    ws4.cell(row=r, column=4, value=cum)
    ws4.cell(row=r, column=5, value=f"{cum/TOTAL_COUPONS*100:.1f}%")
    ws4.cell(row=r, column=6, value=f"${cnt * 3.98:.2f}")
    ws4.cell(row=r, column=7, value=f"${cum * 3.98:,.2f}")
    style_row(ws4, r, len(headers4), fill=fill)

# Summary
sr = 5 + len(coupon_daily) + 1
before_sum = sum(c[1] for c in coupon_daily if c[2] == "活动前")
during_sum = sum(c[1] for c in coupon_daily if c[2] == "活动中")
after_sum = sum(c[1] for c in coupon_daily if c[2] == "活动后")

ws4.cell(row=sr, column=1, value="汇总")
ws4.cell(row=sr, column=1).font = subtitle_font
for j, (label, val) in enumerate([
    ("活动前", before_sum), ("活动中", during_sum), ("活动后", after_sum), ("合计", cum)
]):
    r = sr + 1 + j
    ws4.cell(row=r, column=1, value=label)
    ws4.cell(row=r, column=1).font = bold_font
    ws4.cell(row=r, column=2, value=val)
    ws4.cell(row=r, column=3, value=f"{val/TOTAL_COUPONS*100:.1f}%" if val < TOTAL_COUPONS else "—")
    ws4.cell(row=r, column=6, value=f"${val * 3.98:,.2f}")
    for c in range(1, 8):
        ws4.cell(row=r, column=c).border = thin_border

set_col_widths(ws4, [14, 10, 8, 10, 12, 12, 12])


# ============================================================
# Sheet 5: 结论
# ============================================================
ws5 = wb.create_sheet("结论与建议")
ws5.sheet_properties.tabColor = "EF4444"

ws5.merge_cells("A1:B1")
ws5["A1"] = "Coffee Pass 频次分析 — 核心结论"
ws5["A1"].font = title_font

findings = [
    ("发现1: 券替代而非增量",
     "活动中自然消费杯量从 303 暴跌至 161 (-47%)。\n"
     "用户拿到券后, 把原本要正价买的咖啡改用券消费了, 不是新增需求。"),
    ("发现2: 消费透支效应",
     "活动后自然杯量持续低于基线:\n"
     "后A(02-15~02-23): 136杯 (-55%), 后B: 241杯 (-20%), 后C: 230杯 (-24%)。\n"
     "即使券耗完, 自然消费仍未完全恢复。"),
    ("发现3: 单杯实收未受影响",
     "自然消费单杯实收: $3.91~$4.26, 始终稳定。\n"
     "问题不在价格, 而在频次: 券包提前锁定了消费, 降低了后续购买动力。"),
    ("发现4: 券消费半衰期约2周",
     "活动中 CP 核销 276张 -> 后A 135张 -> 后B 68张 -> 后C 25张。\n"
     "约 2 周消耗一半, 4 周后基本耗完。\n"
     "活动效果的时间窗口很短。"),
    ("发现5: 经济账",
     "券包收入: 121人 x $19.9 = $2,408。\n"
     "自然消费损失: 活动后4期(36天)自然杯量累计少 432 杯 x $4.15 = -$1,793。\n"
     "券消费收入: 595张 x $3.98 = $2,368 (但其中大部分替代了 $4.12 的正价消费)。\n"
     "净效果: 短期现金流回收 vs 中期频次下降的 trade-off。"),
]

r = 3
ws5.column_dimensions["A"].width = 28
ws5.column_dimensions["B"].width = 80

for i, (title, detail) in enumerate(findings):
    ws5.cell(row=r, column=1, value=title)
    ws5.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=11, color="1E293B")
    ws5.cell(row=r + 1, column=2, value=detail)
    ws5.cell(row=r + 1, column=2).font = data_font
    ws5.cell(row=r + 1, column=2).alignment = wrap_align
    ws5.row_dimensions[r + 1].height = 50
    r += 3

# 口径
ws5.cell(row=r, column=1, value="数据口径")
ws5.cell(row=r, column=1).font = subtitle_font
notes = [
    "人群: 121名 Coffee Pass 买券用户 (proposal_no = LKUSCP118713952489488385)",
    "券消费识别: t_order_promotion_detail.promotion_name LIKE 'Coffee Pass%'",
    "订单: DWD dwd_t_ord_order_item_d_inc, order_status=90, order_category='门店订单'",
    "品类: 仅饮品 one_category_name='Drink'",
    "收入: 非CP订单=DWD pay_amount; CP核销杯=$3.98/杯(=$19.9/5)",
    "排除: 测试店铺 NJ Test Kitchen 1&2",
    "期间对齐: 每窗口9天, 日均/总量可直接横向对比",
]
for j, n in enumerate(notes):
    ws5.cell(row=r + 1 + j, column=2, value=n)
    ws5.cell(row=r + 1 + j, column=2).font = note_font


# ============================================================
# Sheet 6: 大盘对比（控制外部因素）
# ============================================================
ws6 = wb.create_sheet("大盘对比")
ws6.sheet_properties.tabColor = "0EA5E9"

ws6.merge_cells("A1:K1")
ws6["A1"] = "Coffee Pass 买券用户 vs 大盘 — 控制外部因素"
ws6["A1"].font = title_font
ws6["A1"].alignment = left_align

ws6.merge_cells("A2:K2")
ws6["A2"] = "大盘 = 同期所有非CP用户 | 人均杯量 = 活跃用户口径 | 指数化: 前9天=100"
ws6["A2"].font = note_font

# --- Part A: 原始数据对比 ---
ws6.cell(row=4, column=1, value="A. 原始数据对比")
ws6.cell(row=4, column=1).font = subtitle_font

headers6a = [
    "期间", "CP活跃", "CP订单", "CP杯量", "CP人均杯量",
    "大盘活跃", "大盘订单", "大盘杯量", "大盘人均杯量"
]
r = 5
for c, h in enumerate(headers6a, 1):
    ws6.cell(row=r, column=c, value=h)
style_row(ws6, r, len(headers6a), fill=header_fill, font=header_font)

mkt_raw = [
    ("前9天", 14423, 23043, 27822),
    ("活动中", 16160, 24603, 29863),
    ("后A(02-15~02-23)", 14229, 20086, 24707),
    ("后B(02-24~03-04)", 17504, 27220, 32687),
    ("后C(03-05~03-13)", 20203, 30476, 36149),
]
cp_raw_totals = [
    ("前9天", 88, 288, 369),
    ("活动中", 107, 399, 448),
    ("后A(02-15~02-23)", 88, 221, 271),
    ("后B(02-24~03-04)", 92, 243, 309),
    ("后C(03-05~03-13)", 74, 202, 255),
]

for i in range(5):
    r = 6 + i
    name = mkt_raw[i][0]
    cp_active, cp_orders, cp_cups = cp_raw_totals[i][1], cp_raw_totals[i][2], cp_raw_totals[i][3]
    mkt_active, mkt_orders, mkt_cups = mkt_raw[i][1], mkt_raw[i][2], mkt_raw[i][3]

    is_hl = (i == 1)
    fill_v = highlight_fill if is_hl else (alt_fill if i % 2 == 0 else None)

    ws6.cell(row=r, column=1, value=name)
    ws6.cell(row=r, column=2, value=cp_active)
    ws6.cell(row=r, column=3, value=cp_orders)
    ws6.cell(row=r, column=4, value=cp_cups)
    ws6.cell(row=r, column=5, value=round(cp_cups / cp_active, 2))
    ws6.cell(row=r, column=6, value=f"{mkt_active:,}")
    ws6.cell(row=r, column=7, value=f"{mkt_orders:,}")
    ws6.cell(row=r, column=8, value=f"{mkt_cups:,}")
    ws6.cell(row=r, column=9, value=round(mkt_cups / mkt_active, 2))
    style_row(ws6, r, len(headers6a), fill=fill_v)
    ws6.cell(row=r, column=1).alignment = left_align

# --- Part B: 指数化对比 ---
r_b = 12
ws6.cell(row=r_b, column=1, value="B. 指数化对比（前9天 = 100）")
ws6.cell(row=r_b, column=1).font = subtitle_font

headers6b = ["期间", "CP指数", "CP变化", "大盘指数", "大盘变化", "超额变化(CP-大盘)", "归因"]
r = r_b + 1
for c, h in enumerate(headers6b, 1):
    ws6.cell(row=r, column=c, value=h)
style_row(ws6, r, len(headers6b), fill=header_fill, font=header_font)

cp_base_avg = cp_raw_totals[0][3] / cp_raw_totals[0][1]  # 369/88 = 4.19
mkt_base_avg = mkt_raw[0][3] / mkt_raw[0][1]  # 27822/14423 = 1.93

for i in range(5):
    r = r_b + 2 + i
    name = mkt_raw[i][0]
    cp_avg = cp_raw_totals[i][3] / cp_raw_totals[i][1]
    mkt_avg = mkt_raw[i][3] / mkt_raw[i][1]
    cp_idx = cp_avg / cp_base_avg * 100
    mkt_idx = mkt_avg / mkt_base_avg * 100
    cp_chg = cp_idx - 100
    mkt_chg = mkt_idx - 100
    excess = cp_chg - mkt_chg

    is_hl = (i == 1)
    fill_v = highlight_fill if is_hl else (alt_fill if i % 2 == 0 else None)

    ws6.cell(row=r, column=1, value=name)
    ws6.cell(row=r, column=2, value=round(cp_idx, 1))
    ws6.cell(row=r, column=3, value="—" if i == 0 else f"{cp_chg:+.1f}")
    ws6.cell(row=r, column=4, value=round(mkt_idx, 1))
    ws6.cell(row=r, column=5, value="—" if i == 0 else f"{mkt_chg:+.1f}")
    ws6.cell(row=r, column=6, value="—" if i == 0 else f"{excess:+.1f}pp")

    if i == 0:
        ws6.cell(row=r, column=7, value="基线")
    elif i == 1:
        ws6.cell(row=r, column=7, value="活动拉升")
        ws6.cell(row=r, column=7).font = pct_green
    else:
        ws6.cell(row=r, column=7, value="CP负面效应")
        ws6.cell(row=r, column=7).font = pct_red

    style_row(ws6, r, len(headers6b), fill=fill_v)
    ws6.cell(row=r, column=1).alignment = left_align

    if i > 0:
        ws6.cell(row=r, column=3).font = pct_f(cp_chg)
        ws6.cell(row=r, column=5).font = pct_f(mkt_chg)
        ws6.cell(row=r, column=6).font = pct_f(excess)

# --- Part C: 结论 ---
r_c = r_b + 2 + 5 + 1
ws6.cell(row=r_c, column=1, value="C. 结论")
ws6.cell(row=r_c, column=1).font = subtitle_font

conclusions = [
    ("大盘也有波动", "后A 大盘人均杯量也下降了 10%（可能是季节/天气因素），但 CP 买家下降了 27%，多出的 17pp 是 CP 特有效应"),
    ("后B/后C 差距扩大", "大盘恢复到接近基线(-3%~-7%)，但 CP 买家仍低 18%~20%。扣除大盘波动后，CP 净负面效应约 -11~-17pp"),
    ("结论", "Coffee Pass 活动导致买券用户频次显著低于大盘趋势，约 -11 到 -17 个百分点的超额下降，确认存在消费透支效应"),
]
for j, (label, text) in enumerate(conclusions):
    r = r_c + 1 + j
    ws6.cell(row=r, column=1, value=label)
    ws6.cell(row=r, column=1).font = bold_font
    ws6.merge_cells(f"B{r}:K{r}")
    ws6.cell(row=r, column=2, value=text)
    ws6.cell(row=r, column=2).font = data_font
    ws6.cell(row=r, column=2).alignment = wrap_align

set_col_widths(ws6, [22, 10, 10, 10, 10, 10, 10, 10, 12, 10, 18])


# ============================================================
# 保存
# ============================================================
output = "/Users/xiaoxiao/Vibe coding/coffee-pass/Coffee_Pass_频次分析_v2.xlsx"
wb.save(output)
print(f"Done: {output}")
