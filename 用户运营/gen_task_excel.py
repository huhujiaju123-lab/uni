#!/usr/bin/env python3
"""生成 US用户运营任务方案_提频任务.xlsx"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "US用户运营任务方案_提频任务.xlsx"
wb = openpyxl.Workbook()

# ── 样式 ──────────────────────────────────────────────
hd_font = Font(bold=True, size=12)
hd_fill = PatternFill("solid", fgColor="1F4E79")
hd_font_w = Font(bold=True, size=11, color="FFFFFF")
sub_fill = PatternFill("solid", fgColor="D6E4F0")
sub_font = Font(bold=True, size=11)
num_fmt = '#,##0'
pct_fmt = '0.0%'
usd_fmt = '$#,##0'
usd2_fmt = '$#,##0.00'
thin = Side(style='thin', color='B0B0B0')
border = Border(top=thin, bottom=thin, left=thin, right=thin)
green_fill = PatternFill("solid", fgColor="E2EFDA")
red_fill = PatternFill("solid", fgColor="FCE4EC")
yellow_fill = PatternFill("solid", fgColor="FFF9C4")

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hd_font_w
        cell.fill = hd_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

def style_row(ws, row, ncols, fill=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if fill:
            cell.fill = fill

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 22)

def write_table(ws, start_row, headers, data, header_style=True):
    """写入表格，返回下一个空行号"""
    r = start_row
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    if header_style:
        style_header(ws, r, len(headers))
    r += 1
    for row_data in data:
        for c, v in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=v)
        style_row(ws, r, len(headers))
        r += 1
    return r + 1

def section_title(ws, row, title, ncols=8):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    return row + 1

# ═══════════════════════════════════════════════════════
# Sheet 1: 频次基线
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "频次基线"

r = section_title(ws1, 1, "一、周频次分布（近4周，2/10-3/9）")
r = write_table(ws1, r,
    ["周频次", "W07", "W08", "W09", "W10", "4周均值", "占比"],
    [
        ["1次", 10157, 9923, 10317, 11648, 10511, "76.2%"],
        ["2次", 2196, 1943, 1824, 2303, 2067, "15.0%"],
        ["3次", 653, 643, 589, 799, 671, "4.9%"],
        ["4次", 284, 266, 212, 352, 279, "2.0%"],
        ["5+次", 152, 196, 154, 254, 189, "1.4%"],
        ["合计", 13442, 12971, 13096, 15356, 13716, "100%"],
    ])

r = section_title(ws1, r, "二、月频次分布（12月/1月/2月）")
r = write_table(ws1, r,
    ["月频次", "2025-12", "2026-01", "2026-02", "占比(2月)"],
    [
        ["1次", 22872, 21536, 22309, "62.6%"],
        ["2次", 6043, 5757, 5986, "16.8%"],
        ["3-4次", 4076, 4078, 4174, "11.7%"],
        ["5-7次", 1715, 1899, 1919, "5.4%"],
        ["8+次", 1005, 1225, 1251, "3.5%"],
        ["合计", 35711, 34495, 35639, "100%"],
    ])

r = section_title(ws1, r, "三、星期分布（2月）")
r = write_table(ws1, r,
    ["星期", "订单数", "用户数", "vs 均值"],
    [
        ["Mon", 14556, 10675, "+0.1%"],
        ["Tue", 15570, 10996, "+7.1%"],
        ["Wed", 17131, 11685, "+17.8%"],
        ["Thu", 16585, 11521, "+14.1%"],
        ["Fri", 15016, 10827, "+3.3%"],
        ["Sat", 12654, 9932, "-12.9%"],
        ["Sun", 11412, 8914, "-21.5%"],
    ])

r = section_title(ws1, r, "四、最长连续消费天数")
r = write_table(ws1, r,
    ["连续天数", "用户数", "占比", "累计占比"],
    [
        ["1天", 55753, "85.4%", "85.4%"],
        ["2天", 6380, "9.8%", "95.2%"],
        ["3天", 1761, "2.7%", "97.9%"],
        ["4-5天", 1100, "1.7%", "99.5%"],
        ["6+天", 261, "0.4%", "100%"],
    ])

auto_width(ws1)

# ═══════════════════════════════════════════════════════
# Sheet 2: 迁移矩阵
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("迁移矩阵")

r = section_title(ws2, 1, "频次迁移矩阵（1月→2月）", 7)
r = write_table(ws2, r,
    ["1月频次", "→2月1次", "→2月2次", "→2月3-4次", "→2月5+次", "→流失", "留存率"],
    [
        ["1次(21,536人)", 2929, 1117, 639, 301, 16550, "23.2%"],
        ["2次(5,757人)", 1353, 715, 534, 292, 2863, "50.3%"],
        ["3-4次(4,078人)", 863, 639, 717, 497, 1362, "66.6%"],
        ["5+次(3,124人)", 329, 401, 633, 1426, 335, "89.3%"],
        ["2月新增", 16835, 3114, 1651, 654, "-", "-"],
    ])

# 高亮流失列
for row_idx in range(3, 8):
    cell = ws2.cell(row=row_idx, column=6)
    if cell.value and isinstance(cell.value, int) and cell.value > 5000:
        cell.fill = red_fill

# 高亮留存率
for row_idx in range(3, 7):
    cell = ws2.cell(row=row_idx, column=7)
    val = cell.value
    if val and "89" in str(val):
        cell.fill = green_fill
    elif val and ("23" in str(val)):
        cell.fill = red_fill

r += 1
ws2.cell(row=r, column=1, value="关键发现：").font = Font(bold=True, size=11)
r += 1
ws2.cell(row=r, column=1, value="• 月1次用户流失率 77%，不干预就丢人")
r += 1
ws2.cell(row=r, column=1, value="• 月5+次用户留存率 89.3%，高频用户无需额外激励")
r += 1
ws2.cell(row=r, column=1, value="• 提频核心：把低频用户从\"试用\"推到\"习惯\"阶段")

auto_width(ws2)

# ═══════════════════════════════════════════════════════
# Sheet 3: 用户分层
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("用户分层")

r = section_title(ws3, 1, "用户分层（近4周 2/10-3/9）", 6)
r = write_table(ws3, r,
    ["频次层", "用户数", "占比", "周均单量", "4周总单量", "任务分配"],
    [
        ["极低频(<0.5/周)", 25011, "63.9%", 0.25, 1.0, "Tier 1"],
        ["低频(0.5-1/周)", 9267, "23.7%", 0.58, 2.3, "Tier 1"],
        ["中低频(1-2/周)", 3645, "9.3%", 1.25, 5.0, "Tier 2"],
        ["中频(2-3/周)", 754, "1.9%", 2.28, 9.1, "Tier 2"],
        ["高频(3+/周)", 486, "1.2%", 4.53, 18.1, "不下发"],
    ])

# 颜色标记
for row_idx in [3, 4]:  # Tier 1
    ws3.cell(row=row_idx, column=6).fill = PatternFill("solid", fgColor="BDD7EE")
for row_idx in [5, 6]:  # Tier 2
    ws3.cell(row=row_idx, column=6).fill = PatternFill("solid", fgColor="F8CBAD")
ws3.cell(row=7, column=6).fill = PatternFill("solid", fgColor="D9D9D9")

r = section_title(ws3, r, "任务分层汇总", 6)
r = write_table(ws3, r,
    ["层级", "覆盖人群", "合计用户", "加权基线(周)", "任务目标", "奖励"],
    [
        ["Tier 1", "极低频 + 低频", 34278, "0.34杯/周", "7天x杯", "50%折扣券"],
        ["Tier 2", "中低频 + 中频", 4399, "1.43杯/周", "7天x杯", "免费饮品券"],
        ["不下发", "高频", 486, "4.53杯/周", "-", "-"],
    ])

auto_width(ws3)

# ═══════════════════════════════════════════════════════
# Sheet 4: 方案测算
# ═══════════════════════════════════════════════════════
ws4 = wb.create_sheet("方案测算")

r = section_title(ws4, 1, "测算假设", 4)
r = write_table(ws4, r,
    ["参数", "值", "依据", ""],
    [
        ["参与率 Tier 1", "5% (≈1,714人)", "保守估计", ""],
        ["参与率 Tier 2", "10% (≈440人)", "活跃用户参与意愿更高", ""],
        ["杯均价", "$4.50", "近期平均实收", ""],
        ["50%券成本", "$1.30/张", "实际券面成本", ""],
        ["免费券成本", "$2.50/张", "实际券面成本", ""],
    ])

r = section_title(ws4, r, "Tier 1 测算（1,714人，基线0.34杯/周）", 4)
r = write_table(ws4, r,
    ["指标", "方案A（目标3杯）", "方案B（目标2杯）", "对比"],
    [
        ["增量杯数", "960", "1,011", "B多5%"],
        ["完成人数", "144 (8.4%)", "405 (23.6%)", "B多181%"],
        ["免费增量占比", "70%", "38%", "A高32pp"],
        ["增量收入", "$4,320", "$4,550", "B多5%"],
        ["奖励成本", "$187", "$527", "A省65%"],
        ["净收入", "$4,133", "$4,023", "A多3%"],
        ["ROI", "23.1x", "8.6x", "A高169%"],
    ])

# 高亮方案A优势行
for row_idx in [r-3, r-4, r-2]:  # 净收入、成本、ROI
    for c in [2]:
        ws4.cell(row=row_idx, column=c).fill = green_fill

r = section_title(ws4, r, "Tier 2 测算（440人，基线1.43杯/周）", 4)
r = write_table(ws4, r,
    ["指标", "方案A（目标5杯）", "方案B（目标4杯）", "对比"],
    [
        ["增量杯数", "866", "889", "B多3%"],
        ["完成人数", "144 (32.7%)", "252 (57.3%)", "B多75%"],
        ["免费增量占比", "63%", "43%", "A高20pp"],
        ["增量收入", "$3,897", "$4,001", "B多3%"],
        ["奖励成本", "$360", "$630", "A省43%"],
        ["净收入", "$3,537", "$3,371", "A多5%"],
        ["ROI", "10.8x", "6.4x", "A高69%"],
    ])

auto_width(ws4)

# ═══════════════════════════════════════════════════════
# Sheet 5: 方案对比
# ═══════════════════════════════════════════════════════
ws5 = wb.create_sheet("方案对比汇总")

r = section_title(ws5, 1, "方案 A vs B 汇总（每周）", 5)
r = write_table(ws5, r,
    ["指标", "方案A（3/5杯）", "方案B（2/4杯）", "差异", "优势方"],
    [
        ["参与用户", "2,154", "2,154", "-", "-"],
        ["总增量杯数", "1,826", "1,900", "+4%", "B"],
        ["总增量收入", "$8,217", "$8,551", "+4%", "B"],
        ["总奖励成本", "$547", "$1,157", "-53%", "★ A"],
        ["净收入", "$7,670", "$7,394", "+4%", "★ A"],
        ["ROI", "15.0x", "7.4x", "+103%", "★ A"],
        ["单杯提频成本", "$0.30", "$0.61", "-51%", "★ A"],
        ["免费增量占比", "67%", "40%", "+27pp", "★ A"],
    ])

# 高亮关键行
for row_idx in range(3, 11):
    cell = ws5.cell(row=row_idx, column=5)
    if cell.value and "A" in str(cell.value):
        for c in range(1, 6):
            ws5.cell(row=row_idx, column=c).fill = green_fill

r += 1
r = section_title(ws5, r, "推荐结论", 5)
r2 = r
ws5.cell(row=r2, column=1, value="推荐方案 A（7天3杯/5杯）").font = Font(bold=True, size=12, color="2E7D32")
r2 += 1
ws5.cell(row=r2, column=1, value="核心理由：方案A的67%增量来自\"免费增量区\"，ROI 15.0x vs 7.4x")
r2 += 1
ws5.cell(row=r2, column=1, value="（用户多买了1-2杯但未达目标，公司无需支付奖励）")
r2 += 1
ws5.cell(row=r2, column=1, value="虽然方案B总杯数略多(+4%)，但奖励成本是方案A的2.1倍，净收入反而更低。")
r2 += 1
ws5.cell(row=r2, column=1, value="")
r2 += 1
ws5.cell(row=r2, column=1, value="一句话：目标高一点 → 完成率低 → 白赚的杯数更多 → ROI更优").font = Font(bold=True, size=11)

r2 += 2
r2 = section_title(ws5, r2, "后续节奏", 5)
ws5.cell(row=r2, column=1, value="W12（本周）").font = Font(bold=True)
ws5.cell(row=r2, column=2, value="确认任务方案 + 产品需求评审")
r2 += 1
ws5.cell(row=r2, column=1, value="W13-14").font = Font(bold=True)
ws5.cell(row=r2, column=2, value="产品开发 + 任务配置 + 资源位设计")
r2 += 1
ws5.cell(row=r2, column=1, value="W15").font = Font(bold=True)
ws5.cell(row=r2, column=2, value="小流量AB测试上线（10%用户）")
r2 += 1
ws5.cell(row=r2, column=1, value="W16").font = Font(bold=True)
ws5.cell(row=r2, column=2, value="评估首期数据 → 迭代参数 → 全量推广")

auto_width(ws5)

# ── 保存 ──────────────────────────────────────────────
wb.save(OUT)
print(f"✅ 已生成 {OUT}")
