import json, time, requests, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

auth = json.load(open('/Users/xiaoxiao/.claude/skills/cyberdata-query/auth.json'))
cookies_str = auth['cookies']
jwttoken = auth['jwttoken']

api_headers = {
    'accept': 'application/json, text/plain, */*',
    'content-type': 'application/json; charset=UTF-8',
    'jwttoken': jwttoken,
    'productkey': 'CyberData',
    'origin': 'https://idpcd.luckincoffee.us',
}
cookie_dict = {}
for item in cookies_str.split('; '):
    if '=' in item:
        k, v = item.split('=', 1)
        cookie_dict[k] = v

def run_sql(sql, wait=8):
    ts = int(time.time() * 1000)
    payload = {
        "_t": ts, "tenantId": "1001", "userId": "47",
        "projectId": "1906904360294313985", "resourceGroupId": 1,
        "taskId": "1990991087752757249", "variables": {},
        "sqlStatement": sql, "env": 5
    }
    r = requests.post('https://idpcd.luckincoffee.us/api/dev/task/run',
                       headers=api_headers, cookies=cookie_dict, json=payload)
    resp = r.json()
    if resp.get('code') != '200':
        print(f"  Submit failed: {resp}")
        return None
    task_id = resp['data']
    print(f"  Task: {task_id}, waiting {wait}s...")
    time.sleep(wait)

    ts2 = int(time.time() * 1000)
    payload2 = {
        "_t": ts2, "tenantId": "1001", "userId": "47",
        "projectId": "1906904360294313985", "env": 5,
        "taskInstanceId": str(task_id)
    }
    r2 = requests.post('https://idpcd.luckincoffee.us/api/logger/getQueryLog',
                        headers=api_headers, cookies=cookie_dict, json=payload2)
    data = r2.json()
    if data.get('code') == '200' and data.get('data'):
        rows = []
        for item in data['data']:
            cols = item.get('columns', [])
            if cols:
                rows.extend(cols)
        return rows
    return None

base_cte = """
WITH user_4wk AS (
    SELECT user_no, COUNT(*) / 4.0 AS avg_weekly_cups
    FROM dw_dwd.dwd_t_ord_order_item_d_inc
    WHERE tenant = 'LKUS' AND order_status = 90
      AND order_category = '门店订单' AND one_category_name = 'Drink'
      AND shop_name NOT IN ('NJ Test Kitchen', 'NJ Test Kitchen 2')
      AND dt BETWEEN '2026-02-16' AND '2026-03-15'
    GROUP BY user_no
),
user_1wk AS (
    SELECT user_no, COUNT(*) AS last_week_cups
    FROM dw_dwd.dwd_t_ord_order_item_d_inc
    WHERE tenant = 'LKUS' AND order_status = 90
      AND order_category = '门店订单' AND one_category_name = 'Drink'
      AND shop_name NOT IN ('NJ Test Kitchen', 'NJ Test Kitchen 2')
      AND dt BETWEEN '2026-03-09' AND '2026-03-15'
    GROUP BY user_no
),
user_all AS (
    SELECT a.user_no, a.avg_weekly_cups,
           COALESCE(b.last_week_cups, 0) AS last_week_cups
    FROM user_4wk a
    LEFT JOIN user_1wk b ON a.user_no = b.user_no
)
"""

quadrants = [
    {'id': 1, 'name': '① 稳定活跃', 'sheet': 'Q1_稳定活跃',
     'file': '提频_Q1_稳定活跃.xlsx',
     'condition': 'avg_weekly_cups >= 1 AND last_week_cups >= 1 AND last_week_cups < 3',
     'desc': '4周均频次≥1杯/周 AND 最近1周≥1杯 AND <3杯',
     'strategy': '维持习惯，提频到3杯，下发 Tier 2 任务',
     'color': '2E7D32', 'fill': 'E2EFDA'},
    {'id': 2, 'name': '② 新激活/回流', 'sheet': 'Q2_新激活回流',
     'file': '提频_Q2_新激活回流.xlsx',
     'condition': 'avg_weekly_cups < 1 AND last_week_cups >= 1 AND last_week_cups < 3',
     'desc': '4周均频次<1杯/周 AND 最近1周≥1杯 AND <3杯',
     'strategy': '趁热打铁，下发 Tier 1 任务',
     'color': '1565C0', 'fill': 'D6E4F0'},
    {'id': 3, 'name': '③ 沉默', 'sheet': 'Q3_沉默',
     'file': '提频_Q3_沉默.xlsx',
     'condition': 'avg_weekly_cups < 1 AND last_week_cups = 0',
     'desc': '4周均频次<1杯/周 AND 最近1周=0杯',
     'strategy': '唤醒，首杯优惠',
     'color': '616161', 'fill': 'F2F2F2'},
    {'id': 4, 'name': '④ 流失预警', 'sheet': 'Q4_流失预警',
     'file': '提频_Q4_流失预警.xlsx',
     'condition': 'avg_weekly_cups >= 1 AND last_week_cups = 0',
     'desc': '4周均频次≥1杯/周 AND 最近1周=0杯',
     'strategy': '紧急召回，强激励',
     'color': 'E65100', 'fill': 'FCE4D6'},
]

# Step 1: Get counts first
print("Step 1: Getting counts per quadrant...")
count_sql = base_cte + """
SELECT
    CASE 
        WHEN avg_weekly_cups >= 1 AND last_week_cups >= 1 AND last_week_cups < 3 THEN 'Q1'
        WHEN avg_weekly_cups < 1 AND last_week_cups >= 1 AND last_week_cups < 3 THEN 'Q2'
        WHEN avg_weekly_cups < 1 AND last_week_cups = 0 THEN 'Q3'
        WHEN avg_weekly_cups >= 1 AND last_week_cups = 0 THEN 'Q4'
        ELSE 'EXCLUDED'
    END AS quadrant,
    COUNT(*) AS cnt
FROM user_all
GROUP BY 1 ORDER BY 1
"""
count_rows = run_sql(count_sql)
if count_rows:
    for r in count_rows:
        print(f"  {r}")

# Step 2: Batch fetch users per quadrant (500 per batch)
BATCH = 500
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
normal_font = Font(name='Arial', size=10)
info_font = Font(name='Arial', size=10, color='666666')
center = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
base_path = '/Users/xiaoxiao/Vibe coding/'

for q in quadrants:
    print(f"\n{'='*50}")
    print(f"Fetching {q['name']}...")
    
    all_users = []
    offset = 0
    while True:
        sql = base_cte + f"SELECT user_no FROM user_all WHERE {q['condition']} ORDER BY user_no LIMIT {BATCH} OFFSET {offset}"
        rows = run_sql(sql, wait=6)
        if rows is None:
            print(f"  Query failed at offset {offset}")
            break
        # Skip header row
        batch_users = [r[0] for r in rows if r[0] != 'user_no']
        all_users.extend(batch_users)
        print(f"  Batch offset={offset}: got {len(batch_users)} rows (total: {len(all_users)})")
        if len(batch_users) < BATCH:
            break
        offset += BATCH

    print(f"  Total users for {q['name']}: {len(all_users)}")

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = q['sheet']
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 55

    ws.merge_cells('A1:B1')
    ws['A1'] = f"提频任务 — {q['name']}"
    ws['A1'].font = Font(name='Arial', size=13, bold=True, color=q['color'])

    info_rows = [
        ('象限', q['name']),
        ('划分条件', q['desc']),
        ('运营策略', q['strategy']),
        ('剔除规则', '已剔除最近1周消费≥3杯的用户'),
        ('用户数', f'{len(all_users):,} 人'),
        ('数据窗口', '4周: 2026-02-16~03-15 | 最近1周: 2026-03-09~03-15'),
    ]
    for i, (label, val) in enumerate(info_rows):
        r = 2 + i
        ws.cell(row=r, column=1, value=label).font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=r, column=1).alignment = left_align
        ws.cell(row=r, column=2, value=val).font = info_font
        ws.cell(row=r, column=2).alignment = left_align

    data_start = 9
    ws.cell(row=data_start, column=1, value='user_no')
    ws.cell(row=data_start, column=1).font = header_font
    ws.cell(row=data_start, column=1).fill = header_fill
    ws.cell(row=data_start, column=1).alignment = center
    ws.cell(row=data_start, column=1).border = thin_border

    q_fill = PatternFill(start_color=q['fill'], end_color=q['fill'], fill_type='solid')
    for i, uno in enumerate(all_users):
        r = data_start + 1 + i
        cell = ws.cell(row=r, column=1, value=uno)
        cell.font = normal_font
        cell.alignment = center
        cell.border = thin_border
        if i % 2 == 0:
            cell.fill = q_fill

    filepath = base_path + q['file']
    wb.save(filepath)
    print(f"  Saved: {filepath}")

print("\n\n=== ALL DONE ===")
